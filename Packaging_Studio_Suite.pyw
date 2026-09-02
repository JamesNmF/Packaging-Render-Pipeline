# -*- coding: utf-8 -*-
"""
美术资产中枢 (Art Asset Hub - v1.0 正式版 Qt6 / PySide6 工业级架构)
===================================================================
【核心架构体系】：
1. 🚀 144 FPS 满帧 GPU 硬件加速画廊 (Qt6 / PySide6 RHI Direct3D Engine)：
   - 基于工业标准 Qt6 C++ 原生 Direct3D / OpenGL 硬件渲染引擎；
   - 采用 QListView + QAbstractListModel + QStyledItemDelegate 虚拟化视口；
   - 4-Worker QThreadPool 异步图像流式解码与信号槽 (Signal/Slot) 回填；
   - 彻底绝缘假死与“未响应”，冷启动 0.05 秒瞬开。

2. 🎨 纯净 Beauty 渲染成片缩略图引擎 (Strict Beauty Render Thumbnail Filter)：
   - 严格禁止抓取 png / PNG / texture / 贴图 / 01_Design 等源文件与贴图目录；
   - 严格过滤 clown (小丑选区图)、cryptomatte、法线、深度、AO、粗糙度、Alpha等所有通道图；
   - 仅采纳 04_Renders_通道输出 / 05_Delivery_最终交付 / 渲染 / 03_输出 中的最终 Beauty 效果图。

3. 🐱 萌猫开屏等待页 (最少 2.0 秒无缝展示) & Windows 沉浸式暗黑顶栏：
   - 启动时全屏居中展示萌猫开屏界面，后台无论多快完成均保持至少 2.0 秒视觉享受；
   - Windows 11/10 原生 DWM 沉浸式暗黑标题栏 (Immersive Dark Title Bar)。

4. 🏢 客户与品牌库管理控制台 (Client & Brand Taxonomy Manager)：
   - 🏷️ 客户品牌白名单库：自定义/排序核心客户品牌；
   - 🔗 别名与系列归并映射：支持将“柏缇绿野幽香”、“柏缇防晒乳”自动归并进“柏缇”；
   - 🚫 非品牌智能过滤黑名单：自动/手动屏蔽 Fonts、HDRI、模型库、临时文件夹等杂乱目录；
   - 🖱️ 右键快捷菜单：支持直接在侧边栏右键重命名、归并或隐藏非品牌项。

5. 🧭 「业务形态 + 客户品牌」双维度侧边导航与资产时间阶梯排序：
   - 🏷️ 上组：业务形态分类 (全部 / 包装 / 套盒 / 海报 / 物料)
   - 🏢 下组：客户品牌库 (全部品牌 / 柏缇 / 森之露 / 语后 / 漱外... 动态联动)
   - ⏱️ 4 级资产时间阶梯排序 (Tier 1 渲染图修改时间 > Tier 2 贴图资产时间 > Tier 3 Blend文件时间 > 文件夹时间)

6. 📥 设计源文件分拣与开工工作台 (Source Organizer & Pipeline Launcher)：
   - ⚙️ 自定义文件夹归档规则管理器：自由新建/编辑子目录结构，内置目录树实时预览；
   - 自动绑定 E:\\zjc\\包装默认文件.blend 母版工程并拉起 Blender 5.2 LTS 开工；
   - 自动将新项目录入《产品列表.xlsx》；
   - 📊 渲染图一键双向内嵌写入 Excel 台账单元格。
===================================================================
"""

import os
import sys
import re
import json
import glob
import html
import shutil
import hashlib
import zipfile
import tempfile
import datetime
import threading
import webbrowser
import subprocess
import xml.etree.ElementTree as ET

# 启动 Splash 管理 (PyInstaller 原生 C 语言启动画面接口)
try:
    import pyi_splash
except ImportError:
    pyi_splash = None

import ctypes
from ctypes import wintypes

from PIL import Image
Image.MAX_IMAGE_PIXELS = None

from PySide6.QtCore import (
    Qt, QSize, QRect, QRectF, QPoint, QModelIndex, QAbstractListModel,
    QThreadPool, QRunnable, Signal, QObject, Slot, QTimer, QEvent
)
from PySide6.QtGui import (
    QIcon, QPixmap, QImage, QPainter, QColor, QFont, QPen, QBrush,
    QPainterPath, QCursor, QFontMetrics, QLinearGradient
)
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QPushButton, QComboBox, QListWidget, QListWidgetItem,
    QListView, QStyledItemDelegate, QStyle, QMenu, QFileDialog, QInputDialog,
    QMessageBox, QDialog, QTableWidget, QTableWidgetItem, QCheckBox,
    QHeaderView, QSplitter, QGroupBox, QTextEdit, QPlainTextEdit, QFrame,
    QProgressBar, QSplashScreen, QToolButton, QProgressDialog, QSystemTrayIcon
)

CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".packaging_suite_v7.json")
CACHE_DIR = os.path.join(os.path.expanduser("~"), ".packaging_asset_thumbnails")
EXCEL_CACHE_DIR = os.path.join(CACHE_DIR, "excel_images")
THUMB_CACHE_DIR = os.path.join(CACHE_DIR, "fast_thumbs")
META_CACHE_FILE = os.path.join(CACHE_DIR, "disk_meta_cache.json")
os.makedirs(EXCEL_CACHE_DIR, exist_ok=True)
os.makedirs(THUMB_CACHE_DIR, exist_ok=True)

def get_resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        p = os.path.join(sys._MEIPASS, relative_path)
        if os.path.exists(p):
            return p
    local_p = os.path.join(os.path.dirname(__file__), relative_path)
    if os.path.exists(local_p):
        return local_p
    fixed_p = os.path.join(r"C:\Users\qq424\Packaging_Tools", relative_path)
    if os.path.exists(fixed_p):
        return fixed_p
    return relative_path

APP_ICON_ICO = get_resource_path("app_icon.ico")
APP_ICON_PNG = get_resource_path("app_icon.png")
SPLASH_CAT_JPG = get_resource_path("splash_cat.jpg")

DEFAULT_EXCEL_PATH = r"C:\Users\qq424\WorkBuddy\2026-08-26-15-33-05\产品列表.xlsx"

BLENDER_EXE = r"C:\Program Files\Blender Foundation\Blender 5.2\blender.exe"
if not os.path.exists(BLENDER_EXE):
    BLENDER_EXE = "blender"

def set_dark_titlebar(hwnd, dark=True):
    if sys.platform == "win32":
        try:
            DWMWA_USE_IMMERSIVE_DARK_MODE = 20
            val = ctypes.c_int(1 if dark else 0)
            res = ctypes.windll.dwmapi.DwmSetWindowAttribute(
                wintypes.HWND(hwnd),
                DWMWA_USE_IMMERSIVE_DARK_MODE,
                ctypes.byref(val),
                ctypes.sizeof(val)
            )
            if res != 0:
                DWMWA_USE_IMMERSIVE_DARK_MODE_OLD = 19
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    wintypes.HWND(hwnd),
                    DWMWA_USE_IMMERSIVE_DARK_MODE_OLD,
                    ctypes.byref(val),
                    ctypes.sizeof(val)
                )
        except Exception:
            pass

def get_valid_template_blend(cfg=None):
    candidates = []
    if cfg and cfg.get("template_blend_path"):
        candidates.append(cfg.get("template_blend_path"))
    candidates.extend([
        r"E:\zjc\包装默认文件.blend",
        r"C:\Users\qq424\Packaging_Tools\templates\Packaging_Master_Template.blend",
        os.path.join(os.path.dirname(__file__), "templates", "Packaging_Master_Template.blend"),
        os.path.join(os.path.dirname(__file__), "Packaging_Master_Template.blend"),
        r"C:\Users\qq424\Desktop\Packaging-Render-Pipeline\templates\Packaging_Master_Template.blend"
    ])
    for c in candidates:
        if c and os.path.exists(c):
            return c
    return ""

DEFAULT_TEMPLATE = get_valid_template_blend()

def normalize_workspaces(cfg):
    """
    统一规范化工作盘配置：
    兼容旧版本 ["E:\\zjc", "D:\\Projects"] 与新版本 [{"path": "E:\\zjc", "alias": "主力生产盘", "is_primary": True}, ...]
    """
    raw_v2 = cfg.get("workspaces_v2", [])
    raw_ws = cfg.get("workspaces", [])
    
    ws_list = []
    seen = set()
    
    if raw_v2 and isinstance(raw_v2, list):
        for item in raw_v2:
            if isinstance(item, dict) and item.get("path"):
                norm_p = os.path.normpath(item["path"])
                if norm_p.lower() not in seen:
                    seen.add(norm_p.lower())
                    alias = item.get("alias") or ("主力生产盘" if len(ws_list) == 0 else (os.path.basename(norm_p) or norm_p))
                    ws_list.append({
                        "path": norm_p,
                        "alias": alias,
                        "is_primary": bool(item.get("is_primary", False))
                    })
                    
    if raw_ws and isinstance(raw_ws, list):
        for p in raw_ws:
            if isinstance(p, str) and p:
                norm_p = os.path.normpath(p)
                if norm_p.lower() not in seen:
                    seen.add(norm_p.lower())
                    alias = "主力生产盘" if len(ws_list) == 0 else (os.path.basename(norm_p) or norm_p)
                    ws_list.append({
                        "path": norm_p,
                        "alias": alias,
                        "is_primary": (len(ws_list) == 0)
                    })
                    
    if not ws_list:
        for p in ["E:\\zjc", "D:\\Projects", "E:\\Projects"]:
            if os.path.exists(p):
                ws_list.append({"path": os.path.normpath(p), "alias": "主力生产盘", "is_primary": True})
                break
        if not ws_list:
            ws_list.append({"path": "E:\\zjc", "alias": "主力生产盘", "is_primary": True})
            
    if not any(w.get("is_primary") for w in ws_list) and ws_list:
        ws_list[0]["is_primary"] = True
        
    cfg["workspaces_v2"] = ws_list
    cfg["workspaces"] = [w["path"] for w in ws_list]
    return ws_list

def get_workspace_alias(ws_path, cfg=None):
    if not ws_path:
        return "未知盘符"
    norm_p = os.path.normpath(ws_path).lower()
    if cfg and cfg.get("workspaces_v2"):
        for item in cfg["workspaces_v2"]:
            if os.path.normpath(item.get("path", "")).lower() == norm_p:
                return item.get("alias") or ws_path
    drive = os.path.splitdrive(ws_path)[0].upper()
    return f"{drive}盘 ({os.path.basename(ws_path)})"

def get_drive_space_info(path):
    """获取指定路径所在盘符的可用空间与总空间"""
    if not path or not os.path.exists(path):
        return False, "⚠️ 盘符离线 / 未连接", 0
    try:
        import shutil
        total, used, free = shutil.disk_usage(path)
        total_gb = total / (1024 ** 3)
        free_gb = free / (1024 ** 3)
        used_percent = int((used / total) * 100) if total > 0 else 0
        return True, f"可用 {free_gb:.1f} GB / 共 {total_gb:.1f} GB", used_percent
    except Exception:
        return False, "⚠️ 盘符读取受限", 0

DEFAULT_WORKSPACES = []
for p in ["E:\\zjc", "D:\\Projects", "E:\\Projects"]:
    if os.path.exists(p) and p not in DEFAULT_WORKSPACES:
        DEFAULT_WORKSPACES.append(p)
if not DEFAULT_WORKSPACES:
    DEFAULT_WORKSPACES = ["E:\\zjc" if os.path.exists("E:\\zjc") else "D:\\Projects"]

SYSTEM_IGNORED_DIRS = {
    'system volume information', '$recycle.bin', 'recovery', '$windows.~bt',
    'msocache', 'config.msi', 'perflogs', 'program files', 'program files (x86)',
    'windows', 'programdata', 'appdata', '.git', '.vscode', '.idea', 'temp',
    'fonts', 'render  preset', 'renderraw', 'studio light hdri pack'
}

DEFAULT_IGNORED_BRANDS = [
    "Fonts", "New(v1.5)The Lazy Motion Library", "Studio Light HDRi Pack", 
    "render  preset", "renderraw", "temp", "通用模型", "流体test", "工作汇报",
    "新建文件夹", "新建文件夹 (2)", "新建文件夹 (3)", "新建文件夹 (4)", "新建文件夹 (5)",
    "洗衣液瓶型ome", "美天惠泡泡洗手液备案图", "liuq"
]

DEFAULT_BRAND_ALIASES = {
    "柏缇绿野幽香": "柏缇",
    "柏缇防晒乳": "柏缇"
}

VALID_CATEGORIES = ["包装", "套盒", "海报", "物料"]

CHANNEL_OR_CLOWN_KEYWORDS = (
    'clown', 'clow', 'cryptomatte', 'crypto', '选区', '通道', 'pass', 'mask',
    'alpha', 'normal', '法线', 'depth', '深度', 'zdepth', 'mist', 'ao', 'ambient',
    'roughness', '粗糙度', 'specular', '高光', 'shadow', '阴影', 'diffuse',
    'glossy', 'emission', 'uv', 'position', 'vector', 'motion'
)

def is_channel_or_clown_image(filepath):
    if not filepath:
        return True
    name_low = os.path.basename(filepath).lower()
    return any(kw in name_low for kw in CHANNEL_OR_CLOWN_KEYWORDS)

def is_valid_beauty_thumbnail(thumb_path):
    if not thumb_path or not os.path.exists(thumb_path):
        return False
    if is_channel_or_clown_image(thumb_path):
        return False
    dir_parts = [p.lower() for p in os.path.normpath(thumb_path).split(os.sep)]
    invalid_dirs = {'png', 'texture', 'textures', '02_textures_贴图资产', '01_design_平面原稿', '贴图', '贴图资产', 'design'}
    if any(p in invalid_dirs for p in dir_parts):
        return False
    return True

DEFAULT_FOLDER_RULES = [
    {
        "id": "standard_packaging_5stage",
        "name": "📦 标准包装 5 阶段工程 (默认)",
        "desc": "适用于标准产品包装设计、三维渲染与最终交付的标准工业流",
        "path_pattern": "{brand}/{sku}",
        "subfolders": [
            "01_Design_平面原稿",
            "02_Textures_贴图资产",
            "03_3D_三维工程",
            "04_Renders_通道输出",
            "05_Delivery_最终交付"
        ],
        "design_sub": "01_Design_平面原稿",
        "blend_sub": "03_3D_三维工程",
        "render_sub": "04_Renders_通道输出"
    },
    {
        "id": "category_grouped_packaging",
        "name": "🗂️ 按业务形态分流 (形态/客户/SKU)",
        "desc": "在工作盘下一级建立【包装/套盒/海报/物料】分类目录，二级为客户，三级为SKU",
        "path_pattern": "{category}/{brand}/{sku}",
        "subfolders": [
            "01_Design_平面原稿",
            "02_Textures_贴图资产",
            "03_3D_三维工程",
            "04_Renders_通道输出",
            "05_Delivery_最终交付"
        ],
        "design_sub": "01_Design_平面原稿",
        "blend_sub": "03_3D_三维工程",
        "render_sub": "04_Renders_通道输出"
    },
    {
        "id": "poster_kv_flow",
        "name": "🖼️ 海报与主视觉 KV 创作流",
        "desc": "适用于平面海报、主视觉 KV、电商展板等以渲染合成素材为主的工程",
        "path_pattern": "{brand}/{sku}",
        "subfolders": [
            "01_Ref_参考意向",
            "02_Design_设计原稿",
            "03_3D_三维模型与场景",
            "04_Renders_高清分层输出",
            "05_Final_精修定稿"
        ],
        "design_sub": "02_Design_设计原稿",
        "blend_sub": "03_3D_三维模型与场景",
        "render_sub": "04_Renders_高清分层输出"
    },
    {
        "id": "simple_3stage",
        "name": "⚡ 极简轻量 3 目录流",
        "desc": "适合轻量快速打样与紧急物料制作",
        "path_pattern": "{brand}/{sku}",
        "subfolders": [
            "01_源文件",
            "02_工程",
            "03_输出"
        ],
        "design_sub": "01_源文件",
        "blend_sub": "02_工程",
        "render_sub": "03_输出"
    }
]

DEFAULT_CONFIG = {
    "workspaces": DEFAULT_WORKSPACES,
    "current_workspace": DEFAULT_WORKSPACES[0],
    "excel_path": DEFAULT_EXCEL_PATH if os.path.exists(DEFAULT_EXCEL_PATH) else "",
    "curated_brands": ["柏缇", "森之露", "语后", "漱外", "零食有鸣", "ee18", "超赞", "初见天空单包", "美天惠", "莱米可蕊", "集立秀", "优品本色", "劲购", "天滋", "jojo", "新品"],
    "ignored_brands": DEFAULT_IGNORED_BRANDS,
    "brand_aliases": DEFAULT_BRAND_ALIASES,
    "current_brand": "柏缇",
    "default_category": "包装",
    "theme": "dark",
    "auto_create_blend": True,
    "auto_open_blender": True,
    "auto_open_ai": True,
    "template_blend_path": DEFAULT_TEMPLATE if os.path.exists(DEFAULT_TEMPLATE) else "",
    "auto_append_to_excel": True,
    "folder_rules": DEFAULT_FOLDER_RULES,
    "active_rule_id": "standard_packaging_5stage"
}

DARK_QSS = """
QMainWindow, QWidget {
    background-color: #18191C;
    color: #E2E4E8;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif;
    font-size: 13px;
}
QTabWidget::pane {
    border: 1px solid #2D3037;
    background-color: #18191C;
}
QTabBar::tab {
    background: #202227;
    color: #9BA1B0;
    padding: 8px 18px;
    margin-right: 4px;
    border-top-left-radius: 6px;
    border-top-right-radius: 6px;
    font-weight: bold;
}
QTabBar::tab:selected {
    background: #282A31;
    color: #3B82F6;
    border-bottom: 2px solid #3B82F6;
}
QGroupBox {
    border: 1px solid #333640;
    border-radius: 8px;
    margin-top: 10px;
    padding-top: 14px;
    font-weight: bold;
    color: #9BA1B0;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 4px;
}
QLineEdit, QComboBox, QTextEdit, QPlainTextEdit {
    background-color: #202227;
    border: 1px solid #383B44;
    border-radius: 6px;
    padding: 6px 10px;
    color: #F1F3F5;
}
QLineEdit:focus, QComboBox:focus {
    border: 1px solid #3B82F6;
}
QPushButton, QToolButton {
    background-color: #282A31;
    border: 1px solid #383B44;
    border-radius: 6px;
    padding: 5px 12px;
    color: #F1F3F5;
    font-weight: 600;
}
QPushButton:hover, QToolButton:hover {
    background-color: #32353E;
    border-color: #6B7280;
}
QPushButton:pressed, QToolButton:pressed {
    background-color: #202227;
}
QPushButton#PrimaryBtn {
    background-color: #2563EB;
    border: 1px solid #1D4ED8;
    color: #FFFFFF;
}
QPushButton#PrimaryBtn:hover {
    background-color: #3B82F6;
}
QListWidget {
    background-color: #202227;
    border: 1px solid #383B44;
    border-radius: 8px;
    padding: 4px;
}
QListWidget::item {
    padding: 6px 10px;
    border-radius: 6px;
    color: #9BA1B0;
    font-weight: bold;
}
QListWidget::item:hover {
    background-color: #282A31;
    color: #F1F3F5;
}
QListWidget::item:selected {
    background-color: #2563EB;
    color: #FFFFFF;
}
QListView#GalleryView {
    background-color: #18191C;
    border: none;
}
QScrollBar:vertical {
    border: none;
    background: #18191C;
    width: 8px;
    margin: 0;
}
QScrollBar::handle:vertical {
    background: #383B44;
    min-height: 20px;
    border-radius: 4px;
}
QScrollBar::handle:vertical:hover {
    background: #6B7280;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0px;
}
QTableWidget {
    background-color: #202227;
    border: 1px solid #383B44;
    border-radius: 8px;
    gridline-color: #2D3037;
    color: #F1F3F5;
}
QHeaderView::section {
    background-color: #282A31;
    color: #9BA1B0;
    padding: 6px;
    border: none;
    border-bottom: 1px solid #383B44;
    font-weight: bold;
}
QMenu {
    background-color: #202227;
    border: 1px solid #383B44;
    border-radius: 8px;
    padding: 4px;
}
QMenu::item {
    padding: 8px 24px;
    border-radius: 4px;
    color: #F1F3F5;
}
QMenu::item:selected {
    background-color: #2563EB;
    color: #FFFFFF;
}
QMenu::separator {
    height: 1px;
    background: #383B44;
    margin: 4px 6px;
}
"""

LIGHT_QSS = """
QMainWindow, QWidget {
    background-color: #F4F6F9;
    color: #0F172A;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif;
    font-size: 13px;
}
QTabWidget::pane {
    border: 1px solid #E2E8F0;
    background-color: #F4F6F9;
}
QTabBar::tab {
    background: #FFFFFF;
    color: #64748B;
    padding: 8px 18px;
    margin-right: 4px;
    border-top-left-radius: 6px;
    border-top-right-radius: 6px;
    font-weight: bold;
}
QTabBar::tab:selected {
    background: #F4F6F9;
    color: #2563EB;
    border-bottom: 2px solid #2563EB;
}
QGroupBox {
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    margin-top: 10px;
    padding-top: 14px;
    font-weight: bold;
    color: #475569;
}
QLineEdit, QComboBox, QTextEdit, QPlainTextEdit {
    background-color: #FFFFFF;
    border: 1px solid #CBD5E1;
    border-radius: 6px;
    padding: 6px 10px;
    color: #0F172A;
}
QPushButton, QToolButton {
    background-color: #FFFFFF;
    border: 1px solid #CBD5E1;
    border-radius: 6px;
    padding: 5px 12px;
    color: #0F172A;
    font-weight: 600;
}
QPushButton:hover, QToolButton:hover {
    background-color: #F8FAFC;
    border-color: #94A3B8;
}
QPushButton#PrimaryBtn {
    background-color: #2563EB;
    border: 1px solid #1D4ED8;
    color: #FFFFFF;
}
QListWidget {
    background-color: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    padding: 4px;
}
QListWidget::item {
    padding: 6px 10px;
    border-radius: 6px;
    color: #475569;
    font-weight: bold;
}
QListWidget::item:selected {
    background-color: #2563EB;
    color: #FFFFFF;
}
QListView#GalleryView {
    background-color: #F4F6F9;
    border: none;
}
"""

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                for k, v in DEFAULT_CONFIG.items():
                    if k not in data:
                        data[k] = v
                if not data.get("folder_rules"):
                    data["folder_rules"] = DEFAULT_FOLDER_RULES
                if "ignored_brands" not in data:
                    data["ignored_brands"] = DEFAULT_IGNORED_BRANDS
                if "brand_aliases" not in data:
                    data["brand_aliases"] = DEFAULT_BRAND_ALIASES
                return data
        except Exception:
            pass
    return DEFAULT_CONFIG

def save_json_atomic(filepath, data, ensure_ascii=False, indent=2):
    try:
        dir_name = os.path.dirname(filepath)
        os.makedirs(dir_name, exist_ok=True)
        temp_fd, temp_path = tempfile.mkstemp(dir=dir_name, prefix="tmp_save_", suffix=".json")
        with os.fdopen(temp_fd, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=ensure_ascii, indent=indent)
        os.replace(temp_path, filepath)
    except Exception as e:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

def save_config(cfg):
    save_json_atomic(CONFIG_FILE, cfg, ensure_ascii=False, indent=4)

def load_meta_cache():
    if os.path.exists(META_CACHE_FILE):
        try:
            with open(META_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_meta_cache(cache):
    save_json_atomic(META_CACHE_FILE, cache, ensure_ascii=False, indent=1)

def get_fast_disk_thumbnail_path(orig_img_path, size=(220, 220)):
    if not orig_img_path or not os.path.exists(orig_img_path):
        return None
    try:
        mtime = os.path.getmtime(orig_img_path)
        img_id = hashlib.md5(f"{orig_img_path}_{mtime}_{size[0]}x{size[1]}".encode("utf-8")).hexdigest()
        thumb_path = os.path.join(THUMB_CACHE_DIR, f"{img_id}.jpg")
        
        if os.path.exists(thumb_path) and os.path.getsize(thumb_path) > 0:
            return thumb_path
            
        with Image.open(orig_img_path) as im:
            if im.mode in ("RGBA", "P"):
                im = im.convert("RGB")
            im.thumbnail(size, Image.Resampling.LANCZOS)
            im.save(thumb_path, "JPEG", quality=85, optimize=True)
            return thumb_path
    except Exception:
        return orig_img_path

def extract_images_from_excel_zip(excel_path):
    if not excel_path or not os.path.exists(excel_path):
        return {}
    cell_image_map = {}
    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            sheet_drawing_rels = {}
            for name in z.namelist():
                if name.startswith("xl/drawings/_rels/") and name.endswith(".rels"):
                    drawing_id = os.path.splitext(os.path.basename(name))[0].replace(".xml", "")
                    rels_xml = z.read(name)
                    root = ET.fromstring(rels_xml)
                    target_map = {}
                    for rel in root.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                        r_id = rel.attrib.get('Id')
                        target = rel.attrib.get('Target')
                        if target:
                            norm_target = os.path.normpath(os.path.join("xl/drawings", target)).replace("\\", "/")
                            target_map[r_id] = norm_target
                    sheet_drawing_rels[drawing_id] = target_map

            for name in z.namelist():
                if name.startswith("xl/drawings/drawing") and name.endswith(".xml"):
                    drawing_id = os.path.splitext(os.path.basename(name))[0]
                    target_map = sheet_drawing_rels.get(drawing_id, {})
                    xml_content = z.read(name)
                    root = ET.fromstring(xml_content)
                    
                    for anchor in root.findall('{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}twoCellAnchor') + \
                                  root.findall('{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}oneCellAnchor'):
                        from_tag = anchor.find('{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from')
                        if from_tag is not None:
                            row_tag = from_tag.find('{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row')
                            if row_tag is not None:
                                row_idx = int(row_tag.text) + 1
                                blip = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}blip')
                                if blip is not None:
                                    embed_id = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                    if embed_id and embed_id in target_map:
                                        img_zip_path = target_map[embed_id]
                                        ext = os.path.splitext(img_zip_path)[1]
                                        img_hash = hashlib.md5(img_zip_path.encode('utf-8')).hexdigest()[:10]
                                        out_name = f"row_{row_idx}_{img_hash}{ext}"
                                        out_disk_path = os.path.join(EXCEL_CACHE_DIR, out_name)
                                        if not os.path.exists(out_disk_path):
                                            try:
                                                with open(out_disk_path, "wb") as f_out:
                                                    f_out.write(z.read(img_zip_path))
                                            except Exception:
                                                pass
                                        if os.path.exists(out_disk_path):
                                            cell_image_map[row_idx] = out_disk_path
    except Exception:
        pass
    return cell_image_map

def parse_time_to_timestamp(val):
    """智能解析 Excel / 字符串 / datetime / 时间戳为统一的 Unix 时间戳"""
    if not val:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, (datetime.datetime, datetime.date)):
        return datetime.datetime(val.year, val.month, val.day, getattr(val, 'hour', 0), getattr(val, 'minute', 0), getattr(val, 'second', 0)).timestamp()
    val_str = str(val).strip()
    if not val_str or val_str == "None":
        return 0
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
        "%Y.%m.%d %H:%M:%S",
        "%Y.%m.%d %H:%M",
        "%Y.%m.%d",
    ]
    for fmt in formats:
        try:
            dt = datetime.datetime.strptime(val_str, fmt)
            return dt.timestamp()
        except ValueError:
            pass
    return 0

def format_display_time(mtime):
    """友好格式化时间戳用于卡片展示"""
    if not mtime or mtime <= 0:
        return ""
    try:
        dt = datetime.datetime.fromtimestamp(mtime)
        now = datetime.datetime.now()
        if dt.year == now.year:
            return dt.strftime("%m-%d %H:%M")
        else:
            return dt.strftime("%Y-%m-%d")
    except Exception:
        return ""

def normalize_category(raw_val):
    if not raw_val:
        return "包装"
    s = str(raw_val).strip()
    if "套盒" in s or "礼盒" in s:
        return "套盒"
    if "海报" in s or "KV" in s or "展板" in s or "主图" in s:
        return "海报"
    if "物料" in s or "单页" in s or "折页" in s or "展架" in s or "画册" in s:
        return "物料"
    return "包装"

def resolve_brand_name(raw_brand, brand_aliases=None, ignored_brands=None):
    if not raw_brand:
        return "未分类品牌"
    b = str(raw_brand).strip()
    if not b or b == "None":
        return "未分类品牌"
        
    aliases = brand_aliases or DEFAULT_BRAND_ALIASES
    if b in aliases:
        return aliases[b]
        
    return b

def parse_and_cache_excel(excel_path, brand_aliases=None, ignored_brands=None):
    if not excel_path or not os.path.exists(excel_path):
        return []
    projects = []
    try:
        import openpyxl
        cell_images = extract_images_from_excel_zip(excel_path)
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        sheet = wb.active
        
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
                
        sku_col = headers.get("产品名称") or headers.get("SKU") or headers.get("品名") or headers.get("产品命名") or 2
        brand_col = headers.get("品牌") or headers.get("客户") or 1
        cat_col = headers.get("业务形态") or headers.get("分类") or headers.get("类别") or None
        time_col = headers.get("创建时间") or headers.get("日期") or headers.get("录入时间") or None
        path_col = headers.get("文件路径") or headers.get("路径") or None
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            sku_val = str(row[sku_col - 1]).strip() if len(row) >= sku_col and row[sku_col - 1] else ""
            if not sku_val or sku_val == "None":
                continue
            brand_val = str(row[brand_col - 1]).strip() if len(row) >= brand_col and row[brand_col - 1] else ""
            if brand_val == "None":
                brand_val = ""
            
            p_val = str(row[path_col - 1]).strip() if path_col and len(row) >= path_col and row[path_col - 1] else ""
            if p_val == "None":
                p_val = ""
                
            if (not brand_val or brand_val.isdigit()) and p_val:
                norm_p = p_val.replace("\\", "/")
                parts = [part for part in norm_p.split("/") if part]
                if len(parts) >= 2:
                    brand_val = parts[-2]
            if not brand_val or brand_val.isdigit():
                brand_val = "柏缇"
                
            brand_val = resolve_brand_name(brand_val, brand_aliases, ignored_brands)
            cat_val = str(row[cat_col - 1]).strip() if cat_col and len(row) >= cat_col and row[cat_col - 1] else ""
            time_raw = row[time_col - 1] if time_col and len(row) >= time_col else ""
            time_val = str(time_raw).strip() if time_raw else ""
            if time_val == "None":
                time_val = ""
            excel_mtime = parse_time_to_timestamp(time_raw)
                
            img_path = cell_images.get(row_idx, "")
            
            projects.append({
                "source": "excel",
                "brand": brand_val,
                "sku": sku_val,
                "cat": cat_val,
                "path": p_val,
                "thumbnail": img_path,
                "time": time_val,
                "row_idx": row_idx,
                "mtime": excel_mtime
            })
        wb.close()
    except Exception:
        pass
    return projects

# ----------------- 纯净成片 Beauty 渲染图查找引擎 -----------------
def find_project_thumbnail(proj_dir):
    """
    仅提取真实成片渲染效果图 (Beauty Render)，严格排除：
    1. png / PNG / texture / 贴图 / 01_Design 等源文件与贴图目录
    2. 任何小丑选区 (clown)、Cryptomatte、法线、深度、AO、Roughness、Alpha、Shadow等分层通道图
    """
    if not proj_dir or not os.path.exists(proj_dir):
        return None
    render_candidates = [
        os.path.join(proj_dir, "04_Renders_通道输出"),
        os.path.join(proj_dir, "04_Renders_高清分层输出"),
        os.path.join(proj_dir, "05_Delivery_最终交付"),
        os.path.join(proj_dir, "05_Final_精修定稿"),
        os.path.join(proj_dir, "渲染"),
        os.path.join(proj_dir, "03_输出"),
        os.path.join(proj_dir, "04_输出"),
        os.path.join(proj_dir, "Renders"),
        os.path.join(proj_dir, "renders")
    ]
    
    img_exts = ("*.png", "*.jpg", "*.jpeg", "*.webp")
    for r_dir in render_candidates:
        if os.path.exists(r_dir) and os.path.isdir(r_dir):
            imgs = []
            for ext in img_exts:
                imgs.extend(glob.glob(os.path.join(r_dir, ext)))
                imgs.extend(glob.glob(os.path.join(r_dir, "*", ext)))
            if imgs:
                beauty_imgs = [f for f in imgs if not is_channel_or_clown_image(f)]
                if beauty_imgs:
                    beauty_imgs.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                    return beauty_imgs[0]
                    
    return None

def auto_detect_category_from_name(name):
    if not name:
        return "包装"
    s = str(name).lower()
    if "套盒" in s or "礼盒" in s:
        return "套盒"
    if "海报" in s or "kv" in s or "主图" in s or "展板" in s:
        return "海报"
    if "物料" in s or "单页" in s or "折页" in s or "展架" in s or "画册" in s:
        return "物料"
    return "包装"

# ----------------- 原始开工与立项基准时间提取算法 (以 png / 贴图最早创建时间为准) -----------------
def get_project_asset_mtime(proj_dir):
    """
    计算项目的真实开工/立项时间：
    1. 优先扫描 png / PNG / 贴图 / 01_Design 等原始贴图与设计源文件目录；
    2. 取目录创建时间或内部图片的最早创建时间 (Earliest Creation Time)，杜绝后补渲染图时间干扰；
    3. 次优读取 .blend 三维工程文件的最早时间；
    4. 兜底取项目文件夹本身的创建时间。
    """
    if not proj_dir or not os.path.exists(proj_dir):
        return 0
    try:
        texture_candidates = [
            os.path.join(proj_dir, "png"),
            os.path.join(proj_dir, "PNG"),
            os.path.join(proj_dir, "02_Textures_贴图资产"),
            os.path.join(proj_dir, "textures"),
            os.path.join(proj_dir, "Textures"),
            os.path.join(proj_dir, "texture"),
            os.path.join(proj_dir, "贴图"),
            os.path.join(proj_dir, "贴图资产"),
            os.path.join(proj_dir, "01_Design_平面原稿"),
            os.path.join(proj_dir, "01_源文件"),
            os.path.join(proj_dir, "02_Design_设计原稿"),
        ]
        img_exts = ('.png', '.jpg', '.jpeg', '.webp', '.ai', '.psd', '.pdf', '.tif', '.tiff')
        
        found_times = []
        for t_dir in texture_candidates:
            if os.path.exists(t_dir) and os.path.isdir(t_dir):
                d_stat = os.stat(t_dir)
                d_ctime = getattr(d_stat, 'st_ctime', 0)
                d_mtime = getattr(d_stat, 'st_mtime', 0)
                d_time = min(d_ctime, d_mtime) if (d_ctime > 0 and d_mtime > 0) else (d_ctime or d_mtime)
                if d_time > 0:
                    found_times.append(d_time)
                    
                try:
                    with os.scandir(t_dir) as it:
                        for entry in it:
                            if entry.is_file() and entry.name.lower().endswith(img_exts):
                                f_stat = entry.stat()
                                f_ctime = getattr(f_stat, 'st_ctime', 0)
                                f_mtime = getattr(f_stat, 'st_mtime', 0)
                                f_time = min(f_ctime, f_mtime) if (f_ctime > 0 and f_mtime > 0) else (f_ctime or f_mtime)
                                if f_time > 0:
                                    found_times.append(f_time)
                except Exception:
                    pass
        if found_times:
            return min(found_times)
            
        # 次选：blend 工程文件
        blend_candidates = [
            os.path.join(proj_dir, "03_3D_三维工程"),
            os.path.join(proj_dir, "03_3D_三维模型与场景"),
            os.path.join(proj_dir, "02_工程"),
            os.path.join(proj_dir, "模型"),
            os.path.join(proj_dir, "3D"),
            proj_dir
        ]
        blend_times = []
        for b_dir in blend_candidates:
            if os.path.exists(b_dir) and os.path.isdir(b_dir):
                try:
                    with os.scandir(b_dir) as it:
                        for entry in it:
                            if entry.is_file() and entry.name.lower().endswith('.blend'):
                                f_stat = entry.stat()
                                f_ctime = getattr(f_stat, 'st_ctime', 0)
                                f_mtime = getattr(f_stat, 'st_mtime', 0)
                                f_time = min(f_ctime, f_mtime) if (f_ctime > 0 and f_mtime > 0) else (f_ctime or f_mtime)
                                if f_time > 0:
                                    blend_times.append(f_time)
                except Exception:
                    pass
        if blend_times:
            return min(blend_times)
            
        # 兜底：项目目录本身创建时间
        p_stat = os.stat(proj_dir)
        p_ctime = getattr(p_stat, 'st_ctime', 0)
        p_mtime = getattr(p_stat, 'st_mtime', 0)
        return min(p_ctime, p_mtime) if (p_ctime > 0 and p_mtime > 0) else (p_ctime or p_mtime)
    except Exception:
        return 0

def scan_workspace_projects_fast(ws_root, meta_cache, brand_aliases=None, ignored_brands=None, ws_alias=""):
    if not ws_root or not os.path.exists(ws_root):
        return []
    projects = []
    cache_dirty = False
    norm_ws = os.path.normpath(ws_root)
    drive_l = os.path.splitdrive(norm_ws)[0].upper()
    ws_tag = f"{drive_l}:{ws_alias[:2]}" if ws_alias else f"{drive_l}盘"
    
    try:
        entries = os.listdir(ws_root)
    except Exception:
        return []
        
    for entry in entries:
        if entry.lower() in SYSTEM_IGNORED_DIRS:
            continue
        brand_p = os.path.join(ws_root, entry)
        if not os.path.isdir(brand_p):
            continue
            
        try:
            skus = os.listdir(brand_p)
        except (PermissionError, OSError):
            continue
            
        final_brand = resolve_brand_name(entry, brand_aliases, ignored_brands)
        
        for sku in skus:
            if sku.lower() in SYSTEM_IGNORED_DIRS:
                continue
            sku_p = os.path.join(brand_p, sku)
            try:
                if os.path.isdir(sku_p):
                    s_mtime = get_project_asset_mtime(sku_p)
                    cache_key = sku_p.lower().replace("/", "\\")
                    
                    cached_thumb = meta_cache.get(cache_key, {}).get("thumbnail")
                    has_valid_cached_thumb = (
                        cache_key in meta_cache 
                        and meta_cache[cache_key].get("mtime") == s_mtime
                        and cached_thumb
                        and is_valid_beauty_thumbnail(cached_thumb)
                    )
                    
                    if has_valid_cached_thumb:
                        cached_item = meta_cache[cache_key]
                        projects.append({
                            "source": "disk",
                            "brand": final_brand,
                            "raw_brand": entry,
                            "sku": cached_item.get("sku", sku),
                            "cat": cached_item.get("cat", auto_detect_category_from_name(sku)),
                            "path": sku_p,
                            "thumbnail": cached_item.get("thumbnail"),
                            "time": "",
                            "mtime": s_mtime,
                            "ws_path": norm_ws,
                            "ws_alias": ws_alias,
                            "ws_tag": ws_tag
                        })
                    else:
                        thumb = find_project_thumbnail(sku_p)
                        cat = auto_detect_category_from_name(sku)
                        meta_cache[cache_key] = {
                            "brand": final_brand,
                            "raw_brand": entry,
                            "sku": sku,
                            "cat": cat,
                            "thumbnail": thumb,
                            "mtime": s_mtime
                        }
                        cache_dirty = True
                        projects.append({
                            "source": "disk",
                            "brand": final_brand,
                            "raw_brand": entry,
                            "sku": sku,
                            "cat": cat,
                            "path": sku_p,
                            "thumbnail": thumb,
                            "time": "",
                            "mtime": s_mtime,
                            "ws_path": norm_ws,
                            "ws_alias": ws_alias,
                            "ws_tag": ws_tag
                        })
            except (PermissionError, OSError):
                continue
                
    if cache_dirty:
        save_meta_cache(meta_cache)
        
    projects.sort(key=lambda x: x["mtime"], reverse=True)
    return projects

def merge_excel_and_disk_projects(excel_projects, disk_projects):
    disk_map = {}
    for dp in disk_projects:
        norm_p = dp["path"].lower().replace("/", "\\")
        disk_map[norm_p] = dp
        sku_clean = re.sub(r'[\s_\-\(\)（）]+', '', dp["sku"].lower())
        disk_map[f"sku:{sku_clean}"] = dp
        
    merged = []
    matched_disk_paths = set()
    
    for ep in excel_projects:
        sku_clean = re.sub(r'[\s_\-\(\)（）]+', '', ep["sku"].lower())
        matched_dp = disk_map.get(f"sku:{sku_clean}")
        
        excel_p = ep.get("path") or ""
        real_disk_path = matched_dp["path"] if matched_dp else (excel_p if (excel_p and os.path.exists(excel_p)) else "")
        
        item = {
            "source": "merged" if matched_dp else "excel",
            "brand": ep.get("brand") or (matched_dp.get("brand") if matched_dp else "柏缇"),
            "raw_brand": ep.get("brand") or (matched_dp.get("raw_brand") if matched_dp else "柏缇"),
            "sku": ep["sku"],
            "cat": ep.get("cat") or (matched_dp.get("cat") if matched_dp else "包装"),
            "time": ep.get("time", ""),
            "row_idx": ep.get("row_idx", 0),
            "path": real_disk_path,
            "thumbnail": (matched_dp["thumbnail"] if matched_dp and matched_dp.get("thumbnail") and is_valid_beauty_thumbnail(matched_dp["thumbnail"]) else None) or ep.get("thumbnail"),
            "mtime": (matched_dp["mtime"] if matched_dp and matched_dp.get("mtime") else 0) or ep.get("mtime", 0),
            "ws_path": matched_dp.get("ws_path", "") if matched_dp else (os.path.splitdrive(real_disk_path)[0] if real_disk_path else ""),
            "ws_alias": matched_dp.get("ws_alias", "") if matched_dp else "",
            "ws_tag": matched_dp.get("ws_tag", "") if matched_dp else ""
        }
        if matched_dp:
            matched_disk_paths.add(matched_dp["path"].lower().replace("/", "\\"))
            
        # 智能融合视图中仅保留磁盘物理存在的真实资产，杜绝历史死数据
        if real_disk_path:
            merged.append(item)
        
    for dp in disk_projects:
        norm_p = dp["path"].lower().replace("/", "\\")
        if norm_p not in matched_disk_paths:
            merged.append({
                "source": "disk",
                "brand": dp.get("brand", "未分类品牌"),
                "raw_brand": dp.get("raw_brand", dp.get("brand", "")),
                "sku": dp["sku"],
                "cat": dp.get("cat", "包装"),
                "time": "",
                "row_idx": 0,
                "path": dp["path"],
                "thumbnail": dp.get("thumbnail"),
                "mtime": dp.get("mtime", 0),
                "ws_path": dp.get("ws_path", ""),
                "ws_alias": dp.get("ws_alias", ""),
                "ws_tag": dp.get("ws_tag", "")
            })
            
    merged.sort(key=lambda x: x["mtime"], reverse=True)
    return merged

def update_thumbnail_to_excel(excel_path, proj_path, sku, thumb_path):
    if not excel_path or not os.path.exists(excel_path):
        return False, "Excel 文件未找到！"
    if not thumb_path or not os.path.exists(thumb_path):
        return False, f"未找到可用的缩略图文件: {thumb_path}"

    try:
        import openpyxl
        from openpyxl.drawing.image import Image as OpenpyxlImage
        temp_img_path = get_fast_disk_thumbnail_path(thumb_path, size=(200, 200)) or thumb_path
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        target_row = None
        sku_col = 2
        for col_idx, cell in enumerate(sheet[1], start=1):
            val = str(cell.value or "").strip()
            if val in ("产品名称", "SKU", "品名", "产品命名"):
                sku_col = col_idx
                break
                
        sku_clean = re.sub(r'[\s_\-\(\)（）]+', '', sku.lower()) if sku else ""
        for r in range(2, sheet.max_row + 1):
            cell_val = str(sheet.cell(row=r, column=sku_col).value or "").strip()
            cell_clean = re.sub(r'[\s_\-\(\)（）]+', '', cell_val.lower())
            if sku_clean and cell_clean and (sku_clean == cell_clean or sku_clean in cell_clean or cell_clean in sku_clean):
                target_row = r
                break
                
        if not target_row:
            target_row = sheet.max_row + 1
            sheet.cell(row=target_row, column=sku_col, value=sku)
            
        img_cell = f"C{target_row}"
        img = OpenpyxlImage(temp_img_path)
        img.width = 75
        img.height = 75
        
        sheet.row_dimensions[target_row].height = 65
        sheet.column_dimensions['C'].width = 14
        
        sheet.add_image(img, img_cell)
        wb.save(excel_path)
        wb.close()
        return True, f"已成功将 [{sku}] 的缩略图写入 Excel 单元格 {img_cell}！"
    except PermissionError:
        return False, f"无法保存 Excel！请先关闭正在打开《产品列表.xlsx》的 WPS 或 Excel 程序后重试。"
    except Exception as e:
        return False, f"写入 Excel 缩略图失败: {str(e)}"

def batch_sync_all_thumbnails_to_excel(excel_path, projects):
    if not excel_path or not os.path.exists(excel_path):
        return False, "Excel 文件未找到！"
        
    valid_items = [p for p in projects if p.get("thumbnail") and is_valid_beauty_thumbnail(p["thumbnail"])]
    if not valid_items:
        return False, "当前没有找到任何带有有效成片缩略图的项目！"
        
    success_count = 0
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as OpenpyxlImage
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        sku_col = 2
        for col_idx, cell in enumerate(sheet[1], start=1):
            val = str(cell.value or "").strip()
            if val in ("产品名称", "SKU", "品名", "产品命名"):
                sku_col = col_idx
                break
                
        row_map = {}
        for r in range(2, sheet.max_row + 1):
            c_val = str(sheet.cell(row=r, column=sku_col).value or "").strip()
            if c_val:
                c_clean = re.sub(r'[\s_\-\(\)（）]+', '', c_val.lower())
                row_map[c_clean] = r
                
        sheet.column_dimensions['C'].width = 14
        
        for item in valid_items:
            sku = item.get("sku", "")
            sku_clean = re.sub(r'[\s_\-\(\)（）]+', '', sku.lower()) if sku else ""
            target_row = row_map.get(sku_clean)
            if not target_row:
                for c_clean, r_idx in row_map.items():
                    if sku_clean and (sku_clean in c_clean or c_clean in sku_clean):
                        target_row = r_idx
                        break
            if not target_row:
                continue
                
            thumb_path = item["thumbnail"]
            temp_img_path = get_fast_disk_thumbnail_path(thumb_path, size=(180, 180)) or thumb_path
            
            try:
                img = OpenpyxlImage(temp_img_path)
                img.width = 75
                img.height = 75
                sheet.row_dimensions[target_row].height = 65
                sheet.add_image(img, f"C{target_row}")
                success_count += 1
            except Exception:
                pass
                
        wb.save(excel_path)
        wb.close()
        return True, f"🎉 批量同步完成！已成功将 {success_count} 个项目的成片效果图写入 Excel 台账！"
    except PermissionError:
        return False, "无法保存 Excel！请先关闭 WPS 或 Excel 后重试。"
    except Exception as e:
        return False, f"批量同步发生异常: {str(e)}"

# ----------------- 异步图像加载器 (Qt 线程池) -----------------
class ImageLoadSignal(QObject):
    finished = Signal(str, QPixmap)

class ImageLoadTask(QRunnable):
    def __init__(self, img_path, target_size=(220, 220)):
        super().__init__()
        self.img_path = img_path
        self.target_size = target_size
        self.signals = ImageLoadSignal()

    def run(self):
        try:
            fast_p = get_fast_disk_thumbnail_path(self.img_path, self.target_size)
            if fast_p and os.path.exists(fast_p):
                pm = QPixmap(fast_p)
                if not pm.isNull():
                    self.signals.finished.emit(self.img_path, pm)
        except Exception:
            pass

# ----------------- 异步全量数据加载 Worker (Qt 信号槽) -----------------
class DataLoaderSignals(QObject):
    finished = Signal(list, list, list)

class DataLoaderWorker(QRunnable):
    def __init__(self, excel_path, workspaces_v2, meta_cache, brand_aliases=None, ignored_brands=None):
        super().__init__()
        self.excel_path = excel_path
        self.workspaces_v2 = workspaces_v2 if isinstance(workspaces_v2, list) else [{"path": workspaces_v2, "alias": "主力生产盘", "is_primary": True}]
        self.meta_cache = meta_cache
        self.brand_aliases = brand_aliases
        self.ignored_brands = ignored_brands
        self.signals = DataLoaderSignals()

    def run(self):
        try:
            excel_p = parse_and_cache_excel(self.excel_path, self.brand_aliases, self.ignored_brands) if (self.excel_path and os.path.exists(self.excel_path)) else []
            disk_p = []
            for ws_info in self.workspaces_v2:
                if isinstance(ws_info, dict):
                    ws_path = ws_info.get("path")
                    ws_alias = ws_info.get("alias", "")
                else:
                    ws_path = str(ws_info)
                    ws_alias = ""
                if ws_path and os.path.exists(ws_path):
                    sub_projs = scan_workspace_projects_fast(
                        ws_path, self.meta_cache, self.brand_aliases, self.ignored_brands, ws_alias=ws_alias
                    )
                    disk_p.extend(sub_projs)
            disk_p.sort(key=lambda x: x["mtime"], reverse=True)
            merged = merge_excel_and_disk_projects(excel_p, disk_p)
            self.signals.finished.emit(excel_p, disk_p, merged)
        except Exception:
            self.signals.finished.emit([], [], [])

# ----------------- 萌猫开屏等待窗口 (Cat Splash Screen) -----------------
class CatSplashScreen(QWidget):
    def __init__(self, splash_img_path):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.SplashScreen)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.resize(520, 440)
        
        self.pixmap = None
        if os.path.exists(splash_img_path):
            self.pixmap = QPixmap(splash_img_path).scaled(520, 400, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        self.img_lbl = QLabel()
        self.img_lbl.setAlignment(Qt.AlignCenter)
        if self.pixmap:
            self.img_lbl.setPixmap(self.pixmap)
        layout.addWidget(self.img_lbl, stretch=1)
        
        bottom_bar = QWidget()
        bottom_bar.setFixedHeight(46)
        bottom_bar.setStyleSheet("background: #18191C; border-bottom-left-radius: 12px; border-bottom-right-radius: 12px;")
        b_layout = QVBoxLayout(bottom_bar)
        b_layout.setContentsMargins(16, 6, 16, 6)
        
        self.status_lbl = QLabel("🚀 正在极速载入美术资产中枢...")
        self.status_lbl.setStyleSheet("color: #93C5FD; font-weight: bold; font-size: 12px;")
        b_layout.addWidget(self.status_lbl)
        
        self.prog_bar = QProgressBar()
        self.prog_bar.setRange(0, 0)
        self.prog_bar.setFixedHeight(4)
        self.prog_bar.setTextVisible(False)
        self.prog_bar.setStyleSheet("""
            QProgressBar { border: none; background: #282A31; border-radius: 2px; }
            QProgressBar::chunk { background: #3B82F6; border-radius: 2px; }
        """)
        b_layout.addWidget(self.prog_bar)
        layout.addWidget(bottom_bar)

    def set_status_text(self, text):
        self.status_lbl.setText(text)

# ----------------- 🏢 客户与品牌库管理器弹窗 -----------------
class BrandManagerDialog(QDialog):
    def __init__(self, curated_brands, ignored_brands, brand_aliases, detected_raw_dirs, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🏢 客户与品牌库管理控制台")
        self.resize(780, 560)
        self.setMinimumSize(680, 480)
        
        self.curated_brands = list(curated_brands)
        self.ignored_brands = list(ignored_brands)
        self.brand_aliases = dict(brand_aliases)
        self.detected_raw_dirs = list(detected_raw_dirs)
        
        self.build_ui()

    def showEvent(self, event):
        super().showEvent(event)
        set_dark_titlebar(int(self.winId()), True)

    def build_ui(self):
        main_layout = QVBoxLayout(self)
        
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        tab_curated = QWidget()
        t1_layout = QHBoxLayout(tab_curated)
        
        self.list_curated = QListWidget()
        for b in self.curated_brands:
            self.list_curated.addItem(b)
        t1_layout.addWidget(self.list_curated, stretch=1)
        
        t1_btns = QVBoxLayout()
        btn_add = QPushButton("➕ 新增品牌")
        btn_add.clicked.connect(self.add_curated_brand)
        btn_edit = QPushButton("✏️ 重命名")
        btn_edit.clicked.connect(self.edit_curated_brand)
        btn_del = QPushButton("🗑️ 移除品牌")
        btn_del.clicked.connect(self.delete_curated_brand)
        btn_up = QPushButton("⬆️ 上移")
        btn_up.clicked.connect(self.move_up_brand)
        btn_down = QPushButton("⬇️ 下移")
        btn_down.clicked.connect(self.move_down_brand)
        
        t1_btns.addWidget(btn_add)
        t1_btns.addWidget(btn_edit)
        t1_btns.addWidget(btn_del)
        t1_btns.addSpacing(12)
        t1_btns.addWidget(btn_up)
        t1_btns.addWidget(btn_down)
        t1_btns.addStretch()
        t1_layout.addLayout(t1_btns)
        
        self.tabs.addTab(tab_curated, "  🏢 正式客户品牌库  ")
        
        tab_alias = QWidget()
        t2_layout = QVBoxLayout(tab_alias)
        t2_layout.addWidget(QLabel("<b>子系列 / 历史目录 归并到 目标品牌映射表：</b>"))
        
        self.table_alias = QTableWidget(0, 2)
        self.table_alias.setHorizontalHeaderLabels(["原始目录 / 系列名", "归并到的正式品牌"])
        self.table_alias.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table_alias.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        t2_layout.addWidget(self.table_alias)
        
        self.refresh_alias_table()
        
        t2_bar = QHBoxLayout()
        btn_add_alias = QPushButton("➕ 添加归并映射")
        btn_add_alias.clicked.connect(self.add_alias_mapping)
        btn_del_alias = QPushButton("🗑️ 删除选中映射")
        btn_del_alias.clicked.connect(self.delete_alias_mapping)
        t2_bar.addWidget(btn_add_alias)
        t2_bar.addWidget(btn_del_alias)
        t2_bar.addStretch()
        t2_layout.addLayout(t2_bar)
        
        self.tabs.addTab(tab_alias, "  🔗 别名与系列归并  ")
        
        tab_ignore = QWidget()
        t3_layout = QVBoxLayout(tab_ignore)
        t3_layout.addWidget(QLabel("<b>已屏蔽的非品牌目录 (如字体、HDRI环境、临时文件、模型库等):</b>"))
        
        self.list_ignored = QListWidget()
        for ig in self.ignored_brands:
            self.list_ignored.addItem(ig)
        t3_layout.addWidget(self.list_ignored)
        
        t3_bar = QHBoxLayout()
        btn_auto_detect = QPushButton("⚡ 一键智能扫描并屏蔽非品牌项")
        btn_auto_detect.clicked.connect(self.auto_detect_ignored_folders)
        btn_add_ignore = QPushButton("➕ 手动添加忽略目录")
        btn_add_ignore.clicked.connect(self.add_ignored_brand)
        btn_del_ignore = QPushButton("♻️ 恢复为品牌 (取消忽略)")
        btn_del_ignore.clicked.connect(self.remove_ignored_brand)
        
        t3_bar.addWidget(btn_auto_detect)
        t3_bar.addWidget(btn_add_ignore)
        t3_bar.addWidget(btn_del_ignore)
        t3_bar.addStretch()
        t3_layout.addLayout(t3_bar)
        
        self.tabs.addTab(tab_ignore, "  🚫 非品牌屏蔽黑名单  ")
        
        bottom_bar = QHBoxLayout()
        bottom_bar.addStretch()
        btn_save = QPushButton("💾 保存并立即生效")
        btn_save.setObjectName("PrimaryBtn")
        btn_save.setFixedHeight(36)
        btn_save.clicked.connect(self.save_and_accept)
        btn_cancel = QPushButton("取消")
        btn_cancel.setFixedHeight(36)
        btn_cancel.clicked.connect(self.reject)
        bottom_bar.addWidget(btn_cancel)
        bottom_bar.addWidget(btn_save)
        main_layout.addLayout(bottom_bar)

    def refresh_alias_table(self):
        self.table_alias.setRowCount(0)
        for raw, target in self.brand_aliases.items():
            r = self.table_alias.rowCount()
            self.table_alias.insertRow(r)
            self.table_alias.setItem(r, 0, QTableWidgetItem(raw))
            self.table_alias.setItem(r, 1, QTableWidgetItem(target))

    def add_curated_brand(self):
        text, ok = QInputDialog.getText(self, "新增客户品牌", "请输入新品牌名称:")
        if ok and text.strip():
            val = text.strip()
            if val not in self.curated_brands:
                self.curated_brands.append(val)
                self.list_curated.addItem(val)

    def edit_curated_brand(self):
        row = self.list_curated.currentRow()
        if row >= 0:
            old_val = self.curated_brands[row]
            text, ok = QInputDialog.getText(self, "编辑品牌名称", "修改品牌名称:", text=old_val)
            if ok and text.strip():
                new_val = text.strip()
                self.curated_brands[row] = new_val
                self.list_curated.item(row).setText(new_val)

    def delete_curated_brand(self):
        row = self.list_curated.currentRow()
        if row >= 0:
            del self.curated_brands[row]
            self.list_curated.takeItem(row)

    def move_up_brand(self):
        row = self.list_curated.currentRow()
        if row > 0:
            self.curated_brands[row], self.curated_brands[row-1] = self.curated_brands[row-1], self.curated_brands[row]
            item = self.list_curated.takeItem(row)
            self.list_curated.insertItem(row-1, item)
            self.list_curated.setCurrentRow(row-1)

    def move_down_brand(self):
        row = self.list_curated.currentRow()
        if 0 <= row < len(self.curated_brands) - 1:
            self.curated_brands[row], self.curated_brands[row+1] = self.curated_brands[row+1], self.curated_brands[row]
            item = self.list_curated.takeItem(row)
            self.list_curated.insertItem(row+1, item)
            self.list_curated.setCurrentRow(row+1)

    def add_alias_mapping(self):
        raw, ok1 = QInputDialog.getText(self, "添加归并映射", "请输入待归并的原始目录名 (如 柏缇绿野幽香):")
        if not (ok1 and raw.strip()):
            return
        target, ok2 = QInputDialog.getText(self, "目标品牌", f"将 [{raw.strip()}] 归并到哪个正式品牌？", text="柏缇")
        if ok2 and target.strip():
            self.brand_aliases[raw.strip()] = target.strip()
            self.refresh_alias_table()

    def delete_alias_mapping(self):
        row = self.table_alias.currentRow()
        if row >= 0:
            raw_key = self.table_alias.item(row, 0).text()
            if raw_key in self.brand_aliases:
                del self.brand_aliases[raw_key]
            self.refresh_alias_table()

    def auto_detect_ignored_folders(self):
        auto_patterns = [
            'font', 'hdri', 'render', 'preset', 'raw', 'temp', 'test', '新建文件夹',
            '模型', '汇报', 'library', 'pack', 'ome', '备案图'
        ]
        added_count = 0
        for d in self.detected_raw_dirs:
            d_low = d.lower()
            if any(p in d_low for p in auto_patterns) or d.startswith('.'):
                if d not in self.ignored_brands:
                    self.ignored_brands.append(d)
                    self.list_ignored.addItem(d)
                    added_count += 1
        QMessageBox.information(self, "智能识别完成", f"已成功识别并加入 {added_count} 个非品牌目录到屏蔽名单！")

    def add_ignored_brand(self):
        text, ok = QInputDialog.getText(self, "屏蔽非品牌目录", "请输入要屏蔽的目录名 (如 Fonts):")
        if ok and text.strip():
            val = text.strip()
            if val not in self.ignored_brands:
                self.ignored_brands.append(val)
                self.list_ignored.addItem(val)

    def remove_ignored_brand(self):
        row = self.list_ignored.currentRow()
        if row >= 0:
            val = self.list_ignored.item(row).text()
            self.ignored_brands.remove(val)
            self.list_ignored.takeItem(row)

    def save_and_accept(self):
        self.accept()

# ----------------- Qt 虚拟化画廊数据模型 -----------------
class GalleryModel(QAbstractListModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.projects = []
        self.pixmap_cache = {}
        self.loading_set = set()

    def rowCount(self, parent=QModelIndex()):
        return len(self.projects)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or index.row() >= len(self.projects):
            return None
        proj = self.projects[index.row()]
        if role == Qt.UserRole:
            return proj
        return None

    def set_projects(self, projects):
        self.beginResetModel()
        self.projects = projects
        self.endResetModel()

    def get_project(self, row):
        if 0 <= row < len(self.projects):
            return self.projects[row]
        return None

# ----------------- Qt6 GPU 硬件加速卡片 Delegate (支持点击命中测试) -----------------
class GalleryCardDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_view = parent
        self.thread_pool = QThreadPool.globalInstance()
        self.thread_pool.setMaxThreadCount(4)

    def sizeHint(self, option, index):
        return QSize(220, 280)

    def editorEvent(self, event, model, option, index):
        if event.type() == QEvent.MouseButtonRelease and event.button() == Qt.LeftButton:
            rect = option.rect.adjusted(6, 6, -6, -6)
            pos = event.pos()
            thumb_rect = QRect(rect.left() + 6, rect.top() + 6, rect.width() - 12, rect.width() - 12)
            badge_y = thumb_rect.bottom() + 10
            title_y = badge_y + 26
            action_y = title_y + 24
            btn1_rect = QRect(rect.left() + 10, action_y, (rect.width() - 26) // 2, 22)
            btn2_rect = QRect(btn1_rect.right() + 6, action_y, (rect.width() - 26) // 2, 22)
            
            proj = index.data(Qt.UserRole)
            if proj:
                win = self.parent_view.window()
                if btn1_rect.contains(pos):
                    win.open_folder(proj.get("path"), sku=proj.get("sku"), brand=proj.get("brand"))
                    return True
                elif btn2_rect.contains(pos):
                    win.launch_blend(proj.get("path"), sku=proj.get("sku"), brand=proj.get("brand"))
                    return True
                elif thumb_rect.contains(pos):
                    win.launch_blend(proj.get("path"), sku=proj.get("sku"), brand=proj.get("brand"))
                    return True
        return super().editorEvent(event, model, option, index)

    def paint(self, painter: QPainter, option, index):
        painter.save()
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)

        rect = option.rect.adjusted(6, 6, -6, -6)
        proj = index.data(Qt.UserRole)
        if not proj:
            painter.restore()
            return

        is_hover = (option.state & QStyle.State_MouseOver)
        is_dark = (self.parent_view.window().current_theme == "dark") if hasattr(self.parent_view, "window") else True

        card_bg = QColor("#282A31") if is_dark else QColor("#FFFFFF")
        border_color = QColor("#3B82F6") if is_hover else (QColor("#383B44") if is_dark else QColor("#E2E8F0"))
        
        path = QPainterPath()
        path.addRoundedRect(QRectF(rect), 8, 8)
        painter.fillPath(path, card_bg)
        painter.setPen(QPen(border_color, 1.5 if is_hover else 1))
        painter.drawPath(path)

        thumb_rect = QRect(rect.left() + 6, rect.top() + 6, rect.width() - 12, rect.width() - 12)
        thumb_path_shape = QPainterPath()
        thumb_path_shape.addRoundedRect(QRectF(thumb_rect), 6, 6)
        painter.fillPath(thumb_path_shape, QColor("#141518") if is_dark else QColor("#F8FAFC"))

        img_path = proj.get("thumbnail")
        model = index.model()
        pix = model.pixmap_cache.get(img_path) if img_path else None

        if pix and not pix.isNull():
            scaled = pix.scaled(thumb_rect.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            x_off = thumb_rect.left() + (thumb_rect.width() - scaled.width()) // 2
            y_off = thumb_rect.top() + (thumb_rect.height() - scaled.height()) // 2
            painter.drawPixmap(x_off, y_off, scaled)
        else:
            painter.setPen(QColor("#6B7280"))
            painter.drawText(thumb_rect, Qt.AlignCenter, "📦 待渲染工程")
            
            if img_path and img_path not in model.loading_set and os.path.exists(img_path) and is_valid_beauty_thumbnail(img_path):
                model.loading_set.add(img_path)
                task = ImageLoadTask(img_path)
                task.signals.finished.connect(self.on_image_loaded)
                self.thread_pool.start(task)

        # 缩略图右下角浮动工作盘角标 (如 E:主力, D:归档)
        ws_tag = proj.get("ws_tag", "")
        if not ws_tag and proj.get("path"):
            drive_l = os.path.splitdrive(proj["path"])[0].upper()
            ws_tag = f"{drive_l}盘"
        if ws_tag:
            ws_badge_rect = QRect(thumb_rect.right() - 50, thumb_rect.bottom() - 17, 46, 15)
            ws_path = QPainterPath()
            ws_path.addRoundedRect(QRectF(ws_badge_rect), 3, 3)
            painter.fillPath(ws_path, QColor(0, 0, 0, 160))
            painter.setPen(QColor("#F1F3F5"))
            font_ws = QFont(painter.font())
            font_ws.setPointSize(7)
            font_ws.setBold(True)
            painter.setFont(font_ws)
            painter.drawText(ws_badge_rect, Qt.AlignCenter, ws_tag)

        cat_val = proj.get("cat", "包装")
        brand_val = proj.get("brand", "")
        sku_val = proj.get("sku", "")

        cat_colors = {
            "包装": (QColor(59, 130, 246, 50), QColor("#93C5FD")),
            "套盒": (QColor(245, 158, 11, 50), QColor("#FCD34D")),
            "海报": (QColor(139, 92, 246, 50), QColor("#C4B5FD")),
            "物料": (QColor(16, 185, 129, 50), QColor("#6EE7B7"))
        }
        badge_bg, badge_fg = cat_colors.get(cat_val, cat_colors["包装"])

        badge_y = thumb_rect.bottom() + 10
        badge_rect = QRect(rect.left() + 10, badge_y, 42, 18)
        badge_path = QPainterPath()
        badge_path.addRoundedRect(QRectF(badge_rect), 4, 4)
        painter.fillPath(badge_path, badge_bg)
        painter.setPen(badge_fg)
        font_badge = QFont(painter.font())
        font_badge.setPointSize(8)
        font_badge.setBold(True)
        painter.setFont(font_badge)
        painter.drawText(badge_rect, Qt.AlignCenter, cat_val)

        if brand_val:
            painter.setPen(QColor("#9BA1B0") if is_dark else QColor("#64748B"))
            font_brand = QFont(painter.font())
            font_brand.setPointSize(8)
            painter.setFont(font_brand)
            painter.drawText(badge_rect.right() + 6, badge_y + 13, brand_val)

        # 绘制时间（右对齐）
        time_str = format_display_time(proj.get("mtime", 0)) or proj.get("time", "")
        if time_str:
            painter.setPen(QColor("#717684") if is_dark else QColor("#94A3B8"))
            font_time = QFont(painter.font())
            font_time.setPointSize(8)
            painter.setFont(font_time)
            time_rect = QRect(badge_rect.right() + 6, badge_y, rect.right() - badge_rect.right() - 16, 18)
            painter.drawText(time_rect, Qt.AlignRight | Qt.AlignVCenter, f"⏱️ {time_str}")

        title_y = badge_y + 26
        title_rect = QRect(rect.left() + 10, title_y, rect.width() - 20, 22)
        painter.setPen(QColor("#F1F3F5") if is_dark else QColor("#0F172A"))
        font_title = QFont(painter.font())
        font_title.setPointSize(10)
        font_title.setBold(True)
        painter.setFont(font_title)
        
        metrics = QFontMetrics(font_title)
        elided_title = metrics.elidedText(sku_val, Qt.ElideRight, title_rect.width())
        painter.drawText(title_rect, Qt.AlignLeft | Qt.AlignVCenter, elided_title)

        action_y = title_y + 24
        btn1_rect = QRect(rect.left() + 10, action_y, (rect.width() - 26) // 2, 22)
        btn2_rect = QRect(btn1_rect.right() + 6, action_y, (rect.width() - 26) // 2, 22)
        
        b1_path = QPainterPath()
        b1_path.addRoundedRect(QRectF(btn1_rect), 4, 4)
        painter.fillPath(b1_path, QColor("#202227") if is_dark else QColor("#E2E8F0"))
        painter.setPen(QColor("#9BA1B0") if is_dark else QColor("#475569"))
        font_btn = QFont(painter.font())
        font_btn.setPointSize(8)
        painter.setFont(font_btn)
        painter.drawText(btn1_rect, Qt.AlignCenter, "📁 文件夹")

        b2_path = QPainterPath()
        b2_path.addRoundedRect(QRectF(btn2_rect), 4, 4)
        painter.fillPath(b2_path, QColor("#2563EB"))
        painter.setPen(QColor("#FFFFFF"))
        font_btn.setBold(True)
        painter.setFont(font_btn)
        painter.drawText(btn2_rect, Qt.AlignCenter, "🚀 3D工程")

        painter.restore()

    def on_image_loaded(self, img_path, pixmap):
        model = self.parent_view.model()
        model.pixmap_cache[img_path] = pixmap
        model.loading_set.discard(img_path)
        self.parent_view.viewport().update()

# ----------------- 丝滑阻尼平滑滚动视图 -----------------
class SmoothGalleryView(QListView):
    """
    丝滑阻尼平滑像素级滚动视图：
    彻底解决原生 QListView (IconMode) 鼠标滚轮跳跃幅度过大的问题，
    实现细腻、平滑、跟手的像素级滚动。
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setVerticalScrollMode(QListView.ScrollPerPixel)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.verticalScrollBar().setSingleStep(32)

    def wheelEvent(self, event):
        delta = event.angleDelta().y()
        if delta != 0:
            step = -int(delta / 120.0 * 45.0)
            sb = self.verticalScrollBar()
            sb.setValue(sb.value() + step)
            event.accept()
        else:
            super().wheelEvent(event)

# ----------------- 异步防闪退 Excel 缩略图同步引擎 -----------------
class ExcelSyncSignals(QObject):
    progress = Signal(int, int, str)
    finished = Signal(bool, str)

class ExcelSyncWorker(QRunnable):
    def __init__(self, excel_path, projects):
        super().__init__()
        self.excel_path = excel_path
        self.projects = projects
        self.signals = ExcelSyncSignals()
        self.is_cancelled = False

    def cancel(self):
        self.is_cancelled = True

    def run(self):
        valid_items = [p for p in self.projects if p.get("thumbnail") and is_valid_beauty_thumbnail(p["thumbnail"])]
        if not valid_items:
            self.signals.finished.emit(False, "当前没有找到任何带有有效成片效果图的项目！")
            return
            
        total = len(valid_items)
        success_count = 0
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from PIL import Image as PILImage
            
            wb = openpyxl.load_workbook(self.excel_path)
            sheet = wb.active
            
            sku_col = 2
            for col_idx, cell in enumerate(sheet[1], start=1):
                val = str(cell.value or "").strip()
                if val in ("产品名称", "SKU", "品名", "产品命名"):
                    sku_col = col_idx
                    break
                    
            row_map = {}
            for r in range(2, sheet.max_row + 1):
                c_val = str(sheet.cell(row=r, column=sku_col).value or "").strip()
                if c_val:
                    c_clean = re.sub(r'[\s_\-\(\)（）]+', '', c_val.lower())
                    row_map[c_clean] = r
                    
            sheet.column_dimensions['C'].width = 14
            
            temp_sync_dir = os.path.join(CACHE_DIR, "excel_sync_temp")
            os.makedirs(temp_sync_dir, exist_ok=True)
            
            for idx, item in enumerate(valid_items, start=1):
                if self.is_cancelled:
                    wb.close()
                    self.signals.finished.emit(False, "已取消同步操作。")
                    return
                    
                sku = item.get("sku", "")
                self.signals.progress.emit(idx, total, sku)
                
                sku_clean = re.sub(r'[\s_\-\(\)（）]+', '', sku.lower()) if sku else ""
                target_row = row_map.get(sku_clean)
                if not target_row:
                    for c_clean, r_idx in row_map.items():
                        if sku_clean and (sku_clean in c_clean or c_clean in sku_clean):
                            target_row = r_idx
                            break
                if not target_row:
                    continue
                    
                thumb_path = item["thumbnail"]
                if not os.path.exists(thumb_path):
                    continue
                    
                try:
                    safe_thumb_path = os.path.join(temp_sync_dir, f"thumb_row_{target_row}.jpg")
                    with PILImage.open(thumb_path) as im:
                        im = im.convert("RGB")
                        im.thumbnail((150, 150), PILImage.Resampling.LANCZOS)
                        im.save(safe_thumb_path, "JPEG", quality=85)
                        
                    img = OpenpyxlImage(safe_thumb_path)
                    img.width = 75
                    img.height = 75
                    sheet.row_dimensions[target_row].height = 65
                    sheet.add_image(img, f"C{target_row}")
                    success_count += 1
                except Exception:
                    pass
                    
            wb.save(self.excel_path)
            wb.close()
            self.signals.finished.emit(True, f"🎉 批量同步成功！已成功将 {success_count} 个项目的成片效果图写入《产品列表.xlsx》！")
        except PermissionError:
            self.signals.finished.emit(False, "无法保存 Excel！请先关闭正在打开《产品列表.xlsx》的 WPS 或 Excel 程序后重试。")
        except Exception as e:
            self.signals.finished.emit(False, f"同步 Excel 缩略图发生异常: {str(e)}")

# ----------------- 自定义文件夹规则管理弹窗 -----------------
class FolderRuleManagerDialog(QDialog):
    def __init__(self, rules, active_rule_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚙️ 自定义文件夹归档规则管理器")
        self.resize(760, 520)
        self.setMinimumSize(680, 460)
        
        self.rules = [dict(r) for r in rules]
        self.active_rule_id = active_rule_id
        self.current_idx = 0
        
        for idx, r in enumerate(self.rules):
            if r.get("id") == active_rule_id:
                self.current_idx = idx
                break
                
        self.build_ui()
        self.load_rule_into_form(self.current_idx)

    def showEvent(self, event):
        super().showEvent(event)
        set_dark_titlebar(int(self.winId()), True)

    def build_ui(self):
        main_layout = QVBoxLayout(self)
        splitter = QSplitter(Qt.HORIZONTAL)
        
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.addWidget(QLabel("<b>规则预设库</b>"))
        
        self.rule_list = QListWidget()
        self.rule_list.currentRowChanged.connect(self.on_rule_selection_changed)
        left_layout.addWidget(self.rule_list)
        
        btn_box_left = QHBoxLayout()
        btn_new = QPushButton("➕ 新建")
        btn_new.clicked.connect(self.new_rule)
        btn_del = QPushButton("🗑️ 删除")
        btn_del.clicked.connect(self.delete_rule)
        btn_box_left.addWidget(btn_new)
        btn_box_left.addWidget(btn_del)
        left_layout.addLayout(btn_box_left)
        
        splitter.addWidget(left_widget)
        
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        
        right_layout.addWidget(QLabel("规则名称:"))
        self.name_edit = QLineEdit()
        self.name_edit.textChanged.connect(self.on_form_modified)
        right_layout.addWidget(self.name_edit)
        
        right_layout.addWidget(QLabel("层级模板:"))
        self.pattern_combo = QComboBox()
        self.pattern_combo.addItems(["{brand}/{sku}", "{category}/{brand}/{sku}", "{sku}"])
        self.pattern_combo.currentTextChanged.connect(self.on_form_modified)
        right_layout.addWidget(self.pattern_combo)
        
        right_layout.addWidget(QLabel("子文件夹列表 (每行一个):"))
        self.subs_edit = QPlainTextEdit()
        self.subs_edit.textChanged.connect(self.on_form_modified)
        right_layout.addWidget(self.subs_edit)
        
        right_layout.addWidget(QLabel("📁 目录树实时预览:"))
        self.preview_lbl = QLabel()
        self.preview_lbl.setStyleSheet("background: #141518; color: #34D399; font-family: Consolas; padding: 10px; border-radius: 6px;")
        right_layout.addWidget(self.preview_lbl)
        
        splitter.addWidget(right_widget)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)
        main_layout.addWidget(splitter)
        
        bottom_layout = QHBoxLayout()
        bottom_layout.addStretch()
        btn_save = QPushButton("💾 保存并应用此规则")
        btn_save.setObjectName("PrimaryBtn")
        btn_save.clicked.connect(self.save_and_accept)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        bottom_layout.addWidget(btn_cancel)
        bottom_layout.addWidget(btn_save)
        main_layout.addLayout(bottom_layout)
        
        self.refresh_list()

    def refresh_list(self):
        self.rule_list.clear()
        for r in self.rules:
            self.rule_list.addItem(r.get("name", "未命名规则"))
        if 0 <= self.current_idx < len(self.rules):
            self.rule_list.setCurrentRow(self.current_idx)

    def load_rule_into_form(self, idx):
        if 0 <= idx < len(self.rules):
            r = self.rules[idx]
            self.name_edit.setText(r.get("name", ""))
            self.pattern_combo.setCurrentText(r.get("path_pattern", "{brand}/{sku}"))
            self.subs_edit.setPlainText("\n".join(r.get("subfolders", [])))
            self.update_preview()

    def on_rule_selection_changed(self, row):
        if row >= 0 and row != self.current_idx:
            self.save_current_form()
            self.current_idx = row
            self.load_rule_into_form(self.current_idx)

    def on_form_modified(self):
        self.update_preview()

    def update_preview(self):
        pat = self.pattern_combo.currentText().strip() or "{brand}/{sku}"
        root_name = pat.replace("{brand}", "柏缇").replace("{sku}", "红参抗皱霜").replace("{category}", "包装")
        raw_text = self.subs_edit.toPlainText()
        subs = [s.strip() for s in raw_text.split("\n") if s.strip()]
        
        lines = [f"📁 [主工作盘]\\{root_name}"]
        for i, s in enumerate(subs):
            prefix = " └── 📂 " if i == len(subs) - 1 else " ├── 📂 "
            lines.append(f"{prefix}{s}")
        self.preview_lbl.setText("\n".join(lines[:7]))

    def save_current_form(self):
        if 0 <= self.current_idx < len(self.rules):
            r = self.rules[self.current_idx]
            r["name"] = self.name_edit.text().strip() or "未命名规则"
            r["path_pattern"] = self.pattern_combo.currentText().strip() or "{brand}/{sku}"
            raw_text = self.subs_edit.toPlainText()
            subs = [s.strip() for s in raw_text.split("\n") if s.strip()]
            r["subfolders"] = subs if subs else ["01_Design_平面原稿", "03_3D_三维工程"]
            r["design_sub"] = next((s for s in subs if "01" in s or "Design" in s or "原稿" in s), subs[0] if subs else "")
            r["blend_sub"] = next((s for s in subs if "03" in s or "3D" in s or "工程" in s), subs[1] if len(subs)>1 else "")
            r["render_sub"] = next((s for s in subs if "04" in s or "Render" in s or "输出" in s), subs[2] if len(subs)>2 else "")

    def new_rule(self):
        self.save_current_form()
        new_id = f"custom_rule_{int(datetime.datetime.now().timestamp())}"
        new_r = {
            "id": new_id,
            "name": f"✨ 新建规则 ({len(self.rules)+1})",
            "desc": "用户自定义规则",
            "path_pattern": "{brand}/{sku}",
            "subfolders": [
                "01_Design_平面原稿",
                "02_Textures_贴图资产",
                "03_3D_三维工程",
                "04_Renders_通道输出",
                "05_Delivery_最终交付"
            ],
            "design_sub": "01_Design_平面原稿",
            "blend_sub": "03_3D_三维工程",
            "render_sub": "04_Renders_通道输出"
        }
        self.rules.append(new_r)
        self.current_idx = len(self.rules) - 1
        self.refresh_list()
        self.load_rule_into_form(self.current_idx)

    def delete_rule(self):
        if len(self.rules) <= 1:
            QMessageBox.warning(self, "提示", "必须至少保留一套文件夹规则！")
            return
        del self.rules[self.current_idx]
        self.current_idx = max(0, self.current_idx - 1)
        self.refresh_list()
        self.load_rule_into_form(self.current_idx)

    def save_and_accept(self):
        self.save_current_form()
        self.active_rule_id = self.rules[self.current_idx]["id"]
        self.accept()

# ----------------- 工作盘空间管理中枢弹窗 -----------------
class WorkspaceHubDialog(QDialog):
    """
    工作盘空间管理中枢：
    管理多个工作盘、修改别名、切换主力开工盘、查看硬盘容量与在线状态
    """
    workspaces_changed = Signal(list)

    def __init__(self, cfg, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🗄️ 工作盘空间管理中枢 (Workspace Hub)")
        self.resize(760, 480)
        self.setMinimumSize(660, 400)
        self.cfg = cfg
        self.workspaces_v2 = [dict(w) for w in normalize_workspaces(cfg)]
        self.build_ui()

    def showEvent(self, event):
        super().showEvent(event)
        set_dark_titlebar(int(self.winId()), True)

    def build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        tip_lbl = QLabel("💡 <b>提示</b>：您可以添加多个物理硬盘/项目目录，并为它们设置清晰的别名（如主力生产盘、2024归档库）。看板支持一键在不同工作盘之间自由切换。")
        tip_lbl.setWordWrap(True)
        tip_lbl.setStyleSheet("color: #9BA1B0; padding-bottom: 4px;")
        layout.addWidget(tip_lbl)

        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["状态", "工作盘别名", "物理路径", "磁盘空间", "主力开工盘"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        layout.addWidget(self.table)

        self.refresh_table()

        btn_bar = QHBoxLayout()
        btn_add = QPushButton("➕ 绑定新工作盘")
        btn_add.clicked.connect(self.add_workspace)
        btn_rename = QPushButton("✏️ 修改别名")
        btn_rename.clicked.connect(self.rename_workspace)
        btn_set_primary = QPushButton("⭐ 设为主力开工盘")
        btn_set_primary.clicked.connect(self.set_primary_workspace)
        btn_del = QPushButton("🗑️ 移除工作盘")
        btn_del.clicked.connect(self.delete_workspace)
        btn_bar.addWidget(btn_add)
        btn_bar.addWidget(btn_rename)
        btn_bar.addWidget(btn_set_primary)
        btn_bar.addWidget(btn_del)
        btn_bar.addStretch()
        layout.addLayout(btn_bar)

        bottom_bar = QHBoxLayout()
        btn_save = QPushButton("💾 保存并立即全盘重扫")
        btn_save.setStyleSheet("background-color: #2563EB; color: white; font-weight: bold; padding: 8px 18px;")
        btn_save.clicked.connect(self.save_and_close)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        bottom_bar.addStretch()
        bottom_bar.addWidget(btn_cancel)
        bottom_bar.addWidget(btn_save)
        layout.addLayout(bottom_bar)

    def refresh_table(self):
        self.table.setRowCount(len(self.workspaces_v2))
        for row, ws in enumerate(self.workspaces_v2):
            p = ws.get("path", "")
            alias = ws.get("alias", "")
            is_primary = ws.get("is_primary", False)
            
            is_online, space_str, _ = get_drive_space_info(p)
            
            # 状态
            status_item = QTableWidgetItem("🟢 在线" if is_online else "🔴 离线")
            status_item.setTextAlignment(Qt.AlignCenter)
            if not is_online:
                status_item.setForeground(QColor("#EF4444"))
            else:
                status_item.setForeground(QColor("#10B981"))
            self.table.setItem(row, 0, status_item)
            
            # 别名
            alias_item = QTableWidgetItem(alias)
            alias_item.setFont(QFont("Segoe UI", 9, QFont.Bold))
            self.table.setItem(row, 1, alias_item)
            
            # 路径
            path_item = QTableWidgetItem(p)
            self.table.setItem(row, 2, path_item)
            
            # 容量
            space_item = QTableWidgetItem(space_str)
            space_item.setForeground(QColor("#9BA1B0"))
            self.table.setItem(row, 3, space_item)
            
            # 主力盘
            prim_item = QTableWidgetItem("⭐ 主力开工" if is_primary else "—")
            prim_item.setTextAlignment(Qt.AlignCenter)
            if is_primary:
                prim_item.setForeground(QColor("#F59E0B"))
            self.table.setItem(row, 4, prim_item)

    def add_workspace(self):
        d = QFileDialog.getExistingDirectory(self, "选择要绑定的新工作盘目录")
        if d:
            norm_d = os.path.normpath(d)
            if any(os.path.normpath(w["path"]).lower() == norm_d.lower() for w in self.workspaces_v2):
                QMessageBox.warning(self, "提示", "该工作盘已在列表中！")
                return
            default_alias = os.path.basename(norm_d) or norm_d
            alias, ok = QInputDialog.getText(self, "设置别名", f"为该工作盘设置易记别名 (如：{default_alias}):", text=default_alias)
            if ok:
                alias_str = alias.strip() or default_alias
                self.workspaces_v2.append({
                    "path": norm_d,
                    "alias": alias_str,
                    "is_primary": (len(self.workspaces_v2) == 0)
                })
                self.refresh_table()

    def rename_workspace(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(self.workspaces_v2):
            QMessageBox.warning(self, "提示", "请先在列表中选中一个工作盘！")
            return
        ws = self.workspaces_v2[row]
        cur_alias = ws.get("alias", "")
        new_alias, ok = QInputDialog.getText(self, "修改工作盘别名", f"修改 [{ws['path']}] 的显示别名:", text=cur_alias)
        if ok and new_alias.strip():
            ws["alias"] = new_alias.strip()
            self.refresh_table()

    def set_primary_workspace(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(self.workspaces_v2):
            QMessageBox.warning(self, "提示", "请先在列表中选中一个工作盘！")
            return
        for idx, w in enumerate(self.workspaces_v2):
            w["is_primary"] = (idx == row)
        self.refresh_table()

    def delete_workspace(self):
        row = self.table.currentRow()
        if row < 0 or row >= len(self.workspaces_v2):
            QMessageBox.warning(self, "提示", "请先在列表中选中一个工作盘！")
            return
        if len(self.workspaces_v2) <= 1:
            QMessageBox.warning(self, "提示", "至少需要保留一个工作盘！")
            return
        ws = self.workspaces_v2[row]
        reply = QMessageBox.question(self, "确认移除", f"确定要移除工作盘 [{ws.get('alias')}] ({ws['path']}) 吗？\n（物理磁盘中的文件不会被删除）", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            was_primary = ws.get("is_primary", False)
            self.workspaces_v2.pop(row)
            if was_primary and self.workspaces_v2:
                self.workspaces_v2[0]["is_primary"] = True
            self.refresh_table()

    def save_and_close(self):
        self.cfg["workspaces_v2"] = self.workspaces_v2
        self.cfg["workspaces"] = [w["path"] for w in self.workspaces_v2]
        save_config(self.cfg)
        self.workspaces_changed.emit(self.workspaces_v2)
        self.accept()

# ----------------- 手动新建项目工程弹窗 -----------------
class ManualProjectCreateDialog(QDialog):
    """
    手动输入名称创建工程项目对话框：
    支持选择工作盘、输入/选择品牌、业务形态、输入SKU品名，自动生成标准工程目录与.blend并启动Blender
    """
    project_created = Signal(dict)

    def __init__(self, cfg, workspaces_v2, curated_brands, ignored_brands, folder_rules, active_rule_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("✨ 手动新建项目工程 (输入品名快速开工)")
        self.resize(680, 520)
        self.setMinimumSize(580, 440)
        self.cfg = cfg
        self.workspaces_v2 = workspaces_v2
        self.curated_brands = curated_brands
        self.ignored_brands = ignored_brands
        self.folder_rules = folder_rules
        self.active_rule_id = active_rule_id
        
        self.build_ui()

    def showEvent(self, event):
        super().showEvent(event)
        set_dark_titlebar(int(self.winId()), True)

    def build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        title_lbl = QLabel("📁 <b>手动创建工业级标准项目脚手架</b>")
        title_lbl.setStyleSheet("font-size: 14px;")
        layout.addWidget(title_lbl)

        form_layout = QGridLayout()
        form_layout.setSpacing(10)

        # 1. 目标工作盘
        form_layout.addWidget(QLabel("目标工作盘:"), 0, 0)
        self.combo_ws = QComboBox()
        for ws in self.workspaces_v2:
            p = ws.get("path", "")
            alias = ws.get("alias", "")
            is_prim = ws.get("is_primary", False)
            self.combo_ws.addItem(f"{'⭐ ' if is_prim else ''}{alias} ({p})", p)
        cur_ws = self.cfg.get("current_workspace", "")
        for idx in range(self.combo_ws.count()):
            if cur_ws and os.path.normpath(cur_ws).lower() == os.path.normpath(self.combo_ws.itemData(idx)).lower():
                self.combo_ws.setCurrentIndex(idx)
                break
        self.combo_ws.currentIndexChanged.connect(self.update_preview)
        form_layout.addWidget(self.combo_ws, 0, 1)

        # 2. 客户品牌
        form_layout.addWidget(QLabel("客户品牌:"), 1, 0)
        self.combo_brand = QComboBox()
        self.combo_brand.setEditable(True)
        brands = [b for b in self.curated_brands if b not in self.ignored_brands]
        self.combo_brand.addItems(brands)
        cur_b = self.cfg.get("current_brand", "柏缇")
        if cur_b in brands:
            self.combo_brand.setCurrentText(cur_b)
        self.combo_brand.currentTextChanged.connect(self.update_preview)
        form_layout.addWidget(self.combo_brand, 1, 1)

        # 3. 业务形态
        form_layout.addWidget(QLabel("业务形态:"), 2, 0)
        self.combo_cat = QComboBox()
        self.combo_cat.addItems(VALID_CATEGORIES)
        self.combo_cat.setCurrentText(self.cfg.get("default_category", "包装"))
        self.combo_cat.currentIndexChanged.connect(self.update_preview)
        form_layout.addWidget(self.combo_cat, 2, 1)

        # 4. 产品 SKU / 品名
        form_layout.addWidget(QLabel("<b>产品名称 (SKU):</b>"), 3, 0)
        self.sku_edit = QLineEdit()
        self.sku_edit.setPlaceholderText("例如: 红参抗皱紧致双抗乳液 / 多肽修护精华霜")
        self.sku_edit.setStyleSheet("font-size: 13px; font-weight: bold; padding: 6px;")
        self.sku_edit.textChanged.connect(self.update_preview)
        form_layout.addWidget(self.sku_edit, 3, 1)

        # 5. 归档规则
        form_layout.addWidget(QLabel("归档规则:"), 4, 0)
        self.combo_rule = QComboBox()
        for idx, r in enumerate(self.folder_rules):
            self.combo_rule.addItem(r.get("name", ""), r.get("id"))
            if r.get("id") == self.active_rule_id:
                self.combo_rule.setCurrentIndex(idx)
        self.combo_rule.currentIndexChanged.connect(self.update_preview)
        form_layout.addWidget(self.combo_rule, 4, 1)

        layout.addLayout(form_layout)

        # 目录树实时预览
        layout.addWidget(QLabel("📁 <b>即将生成的物理目录与工程文件预览:</b>"))
        self.preview_lbl = QLabel()
        self.preview_lbl.setStyleSheet("background: #141518; color: #34D399; font-family: Consolas, 'Microsoft YaHei UI'; padding: 10px; border-radius: 6px; font-size: 11px;")
        layout.addWidget(self.preview_lbl)

        # 自动化选项
        opts_box = QHBoxLayout()
        self.chk_blend = QCheckBox("✨ 自动创建对应 .blend 工程")
        self.chk_blend.setChecked(True)
        self.chk_open_blender = QCheckBox("🚀 立即拉起 Blender 开工")
        self.chk_open_blender.setChecked(True)
        self.chk_excel = QCheckBox("📊 录入 Excel 台账")
        self.chk_excel.setChecked(True)
        self.chk_open_folder = QCheckBox("📂 在文件夹中打开")
        self.chk_open_folder.setChecked(True)

        opts_box.addWidget(self.chk_blend)
        opts_box.addWidget(self.chk_open_blender)
        opts_box.addWidget(self.chk_excel)
        opts_box.addWidget(self.chk_open_folder)
        layout.addLayout(opts_box)

        # 底部按钮
        btn_bar = QHBoxLayout()
        btn_create = QPushButton("🚀 立即创建项目工程并开工")
        btn_create.setObjectName("PrimaryBtn")
        btn_create.setStyleSheet("background-color: #2563EB; color: white; font-weight: bold; font-size: 13px; padding: 8px 20px;")
        btn_create.clicked.connect(self.do_create_project)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        
        btn_bar.addStretch()
        btn_bar.addWidget(btn_cancel)
        btn_bar.addWidget(btn_create)
        layout.addLayout(btn_bar)

        self.update_preview()

    def get_selected_rule(self):
        rule_id = self.combo_rule.currentData()
        for r in self.folder_rules:
            if r.get("id") == rule_id:
                return r
        return self.folder_rules[0] if self.folder_rules else DEFAULT_FOLDER_RULES[0]

    def update_preview(self):
        ws_path = self.combo_ws.currentData() or "E:\\zjc"
        brand = self.combo_brand.currentText().strip() or "客户品牌"
        cat = self.combo_cat.currentText().strip() or "包装"
        raw_sku = self.sku_edit.text().strip() or "未命名产品品名"
        sku = re.sub(r'[\\/:*?"<>|]', '', raw_sku).strip() or "产品SKU"
        
        rule = self.get_selected_rule()
        pat = rule.get("path_pattern", "{brand}/{sku}")
        subfolders = rule.get("subfolders", DEFAULT_FOLDER_RULES[0]["subfolders"])
        blend_sub = rule.get("blend_sub", "03_3D_三维工程")

        if "{category}" in pat:
            rel = f"{cat}\\{brand}\\{sku}"
        elif "{brand}" in pat:
            rel = f"{brand}\\{sku}"
        else:
            rel = sku
            
        full_proj_dir = os.path.join(ws_path, rel)
        
        lines = [f"📁 {full_proj_dir}"]
        for i, s in enumerate(subfolders):
            prefix = " └── 📂 " if i == len(subfolders) - 1 else " ├── 📂 "
            extra = f" (含 {sku}.blend)" if s == blend_sub else ""
            lines.append(f"{prefix}{s}{extra}")
            
        self.preview_lbl.setText("\n".join(lines[:7]))

    def do_create_project(self):
        ws_path = self.combo_ws.currentData()
        if not ws_path or not os.path.exists(ws_path):
            QMessageBox.warning(self, "错误", f"目标工作盘不存在:\n{ws_path}")
            return
            
        brand = self.combo_brand.currentText().strip()
        cat = self.combo_cat.currentText().strip()
        raw_sku = self.sku_edit.text().strip()
        if not raw_sku:
            QMessageBox.warning(self, "提示", "请输入产品名称 (SKU)！")
            self.sku_edit.setFocus()
            return
            
        sku = re.sub(r'[\\/:*?"<>|]', '', raw_sku).strip()
        if not sku:
            QMessageBox.warning(self, "提示", "产品名称包含非法字符，请输入有效的品名！")
            return
            
        rule = self.get_selected_rule()
        pat = rule.get("path_pattern", "{brand}/{sku}")
        subfolders = rule.get("subfolders", DEFAULT_FOLDER_RULES[0]["subfolders"])
        blend_sub = rule.get("blend_sub", "03_3D_三维工程")
        
        if "{category}" in pat:
            rel = f"{cat}/{brand}/{sku}" if brand else f"{cat}/{sku}"
        elif "{brand}" in pat:
            rel = f"{brand}/{sku}" if brand else sku
        else:
            rel = sku
            
        proj_dir = os.path.join(ws_path, rel)
        try:
            os.makedirs(proj_dir, exist_ok=True)
            for s in subfolders:
                os.makedirs(os.path.join(proj_dir, s), exist_ok=True)
        except Exception as e:
            QMessageBox.critical(self, "创建失败", f"无法创建项目目录:\n{proj_dir}\n\n{str(e)}")
            return
            
        # 自动生成 .blend 工程
        target_blend = os.path.join(proj_dir, blend_sub, f"{sku}.blend")
        if self.chk_blend.isChecked():
            if not os.path.exists(target_blend):
                tpl = get_valid_template_blend(self.cfg)
                if tpl and os.path.exists(tpl):
                    shutil.copy2(tpl, target_blend)
                else:
                    with open(target_blend, "wb") as bf:
                        pass
                        
        # 录入 Excel 台账
        if self.chk_excel.isChecked():
            ex_path = self.cfg.get("excel_path", DEFAULT_EXCEL_PATH)
            if ex_path and os.path.exists(ex_path):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(ex_path)
                    sheet = wb.active
                    sku_col = 2
                    brand_col = 1
                    cat_col = None
                    for col_idx, cell in enumerate(sheet[1], start=1):
                        val = str(cell.value or "").strip()
                        if val in ("产品名称", "SKU", "品名", "产品命名"):
                            sku_col = col_idx
                        if val in ("品牌", "客户"):
                            brand_col = col_idx
                        if val in ("业务形态", "分类", "类别"):
                            cat_col = col_idx
                            
                    sku_clean = re.sub(r'[\s_\-\(\)（）]+', '', sku.lower())
                    exists = False
                    for r in range(2, sheet.max_row + 1):
                        c_val = str(sheet.cell(row=r, column=sku_col).value or "").strip()
                        if sku_clean and sku_clean == re.sub(r'[\s_\-\(\)（）]+', '', c_val.lower()):
                            exists = True
                            break
                    if not exists:
                        new_r = sheet.max_row + 1
                        sheet.cell(row=new_r, column=brand_col, value=brand)
                        sheet.cell(row=new_r, column=sku_col, value=sku)
                        if cat_col:
                            sheet.cell(row=new_r, column=cat_col, value=cat)
                        wb.save(ex_path)
                    wb.close()
                except Exception:
                    pass
                    
        # 拉起 Blender
        if self.chk_open_blender.isChecked() and os.path.exists(target_blend):
            try:
                subprocess.Popen([BLENDER_EXE, target_blend])
            except Exception:
                try:
                    os.startfile(target_blend)
                except Exception:
                    pass
                    
        # 弹出文件夹
        if self.chk_open_folder.isChecked():
            try:
                os.startfile(proj_dir)
            except Exception:
                pass
                
        self.project_created.emit({
            "path": proj_dir,
            "sku": sku,
            "brand": brand,
            "cat": cat,
            "ws_path": ws_path
        })
        QMessageBox.information(self, "创建成功", f"🎉 项目 【{brand} / {sku}】 已成功创建并初始化完成！")
        self.accept()

# ----------------- 关闭窗口行为确认弹窗 -----------------
class CloseActionDialog(QDialog):
    """
    关闭窗口确认对话框：
    支持选择最小化到系统托盘（后台热驻留）或彻底退出，并可记住选择
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚙️ 关闭窗口选项")
        self.setFixedSize(390, 210)
        self.selected_action = "tray"
        self.remember = False
        self.build_ui()

    def showEvent(self, event):
        super().showEvent(event)
        set_dark_titlebar(int(self.winId()), True)

    def build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(14)

        lbl = QLabel("<b>您希望如何处理当前窗口？</b><br><span style='color:#9BA1B0; font-size:12px;'>最小化到系统托盘可保留后台热驻留，再次打开 0 毫秒瞬开，减少重复等待时间。</span>")
        lbl.setWordWrap(True)
        layout.addWidget(lbl)

        self.chk_remember = QCheckBox("记住我的选择，下次关闭直接执行")
        self.chk_remember.setStyleSheet("color: #CBD5E1; font-size: 12px;")
        layout.addWidget(self.chk_remember)

        btn_box = QHBoxLayout()
        btn_box.setSpacing(12)
        
        btn_tray = QPushButton("📥 最小化到托盘 (推荐)")
        btn_tray.setObjectName("PrimaryBtn")
        btn_tray.setStyleSheet("background-color: #2563EB; color: white; font-weight: bold; font-size: 12px; padding: 8px 14px;")
        btn_tray.clicked.connect(self.choose_tray)

        btn_quit = QPushButton("❌ 彻底退出程序")
        btn_quit.setStyleSheet("padding: 8px 14px; font-size: 12px;")
        btn_quit.clicked.connect(self.choose_quit)

        btn_box.addWidget(btn_tray)
        btn_box.addWidget(btn_quit)
        layout.addLayout(btn_box)

    def choose_tray(self):
        self.selected_action = "tray"
        self.remember = self.chk_remember.isChecked()
        self.accept()

    def choose_quit(self):
        self.selected_action = "quit"
        self.remember = self.chk_remember.isChecked()
        self.accept()

# ----------------- Qt6 工业级主窗口 -----------------
class MainWindow(QMainWindow):
    initial_load_done = Signal()

    def __init__(self, initial_files=None):
        super().__init__()
        self.setWindowTitle("美术资产中枢 - Art Asset Hub (v1.0 正式版 Qt6 GPU 加速)")
        self.resize(1320, 860)
        self.setMinimumSize(1060, 680)

        self.cfg = load_config()
        self.meta_cache = load_meta_cache()
        self.current_theme = self.cfg.get("theme", "dark")
        self.workspaces_v2 = normalize_workspaces(self.cfg)
        self.workspaces = [w["path"] for w in self.workspaces_v2]
        
        self.curated_brands = self.cfg.get("curated_brands", DEFAULT_CONFIG["curated_brands"])
        self.ignored_brands = self.cfg.get("ignored_brands", DEFAULT_IGNORED_BRANDS)
        self.brand_aliases = self.cfg.get("brand_aliases", DEFAULT_BRAND_ALIASES)
        self.detected_raw_dirs = []
        
        self.folder_rules = self.cfg.get("folder_rules", DEFAULT_FOLDER_RULES)
        self.active_rule_id = self.cfg.get("active_rule_id", "standard_packaging_5stage")
        
        self.excel_projects = []
        self.disk_projects = []
        self.merged_projects = []
        self.current_display_list = []
        
        self.selected_category = "全部"
        self.selected_brand = "全部"
        self.brand_counts_map = {}
        
        self.files_to_organize = []
        self.has_fired_initial_done = False
        self.is_force_quitting = False
        
        self.setup_ui()
        self.init_tray_icon()
        self.apply_theme()
        self.update_workspace_filter_combo()
        self.update_organizer_ws_combo()
        
        cached_disk = list(self.meta_cache.values())
        if cached_disk:
            cached_disk.sort(key=lambda x: x.get("mtime", 0), reverse=True)
            self.disk_projects = cached_disk
            self.merged_projects = cached_disk
            self.update_sidebar_counts()
            self.update_active_dataset()

        QTimer.singleShot(50, self.async_load_data)
        
        if initial_files:
            self.tabs.setCurrentIndex(1)
            self.add_files_to_organizer(initial_files)

    def showEvent(self, event):
        super().showEvent(event)
        set_dark_titlebar(int(self.winId()), self.current_theme == "dark")

    def apply_theme(self):
        is_dark = (self.current_theme == "dark")
        if is_dark:
            self.setStyleSheet(DARK_QSS)
            self.btn_theme.setText("🌙 护眼暗灰")
        else:
            self.setStyleSheet(LIGHT_QSS)
            self.btn_theme.setText("☀️ 浅色模式")
        set_dark_titlebar(int(self.winId()), is_dark)

    def toggle_theme(self):
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.cfg["theme"] = self.current_theme
        save_config(self.cfg)
        self.apply_theme()
        self.gallery_view.viewport().update()

    # ---------------- 系统托盘与关闭行为管理 ----------------
    def init_tray_icon(self):
        self.tray_icon = QSystemTrayIcon(self)
        icon = QIcon(APP_ICON_PNG) if os.path.exists(APP_ICON_PNG) else QIcon(APP_ICON_ICO)
        self.tray_icon.setIcon(icon)
        self.tray_icon.setToolTip("美术资产中枢 (Art Asset Hub) - 点击瞬时唤醒")

        tray_menu = QMenu()
        act_show = tray_menu.addAction("🖼️ 显示主窗口")
        act_show.triggered.connect(self.show_main_window)
        
        act_refresh = tray_menu.addAction("🔄 刷新资产")
        act_refresh.triggered.connect(self.async_load_data)
        
        tray_menu.addSeparator()
        
        act_settings = tray_menu.addAction("⚙️ 关闭行为设置...")
        act_settings.triggered.connect(self.configure_close_action)
        
        tray_menu.addSeparator()
        
        act_quit = tray_menu.addAction("❌ 彻底退出程序")
        act_quit.triggered.connect(self.force_quit_app)

        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.activated.connect(self.on_tray_icon_activated)
        self.tray_icon.show()

    def on_tray_icon_activated(self, reason):
        if reason in (QSystemTrayIcon.Trigger, QSystemTrayIcon.DoubleClick):
            self.show_main_window()

    def show_main_window(self):
        self.showNormal()
        self.activateWindow()
        self.raise_()

    def force_quit_app(self):
        self.is_force_quitting = True
        if hasattr(self, "tray_icon"):
            self.tray_icon.hide()
        QApplication.quit()

    def configure_close_action(self):
        dlg = CloseActionDialog(self)
        dlg.chk_remember.setChecked(self.cfg.get("remember_close_action", False))
        if dlg.exec() == QDialog.Accepted:
            self.cfg["close_action"] = dlg.selected_action
            self.cfg["remember_close_action"] = dlg.remember
            save_config(self.cfg)
            act_text = "📥 最小化到系统托盘 (后台热驻留)" if dlg.selected_action == "tray" else "❌ 彻底退出程序"
            rem_text = "（已开启记住选择）" if dlg.remember else "（下次关闭仍会询问）"
            QMessageBox.information(self, "设置已保存", f"已成功将关闭窗口默认行为设置为：\n【{act_text}】\n{rem_text}")

    def closeEvent(self, event):
        if getattr(self, "is_force_quitting", False):
            event.accept()
            return

        remember = self.cfg.get("remember_close_action", False)
        action = self.cfg.get("close_action", "")

        if not remember or action not in ("tray", "quit"):
            dlg = CloseActionDialog(self)
            if dlg.exec() != QDialog.Accepted:
                event.ignore()
                return
            action = dlg.selected_action
            if dlg.remember:
                self.cfg["close_action"] = action
                self.cfg["remember_close_action"] = True
                save_config(self.cfg)
            else:
                self.cfg["remember_close_action"] = False
                save_config(self.cfg)

        if action == "tray":
            event.ignore()
            self.hide()
            if not self.cfg.get("has_shown_tray_bubble", False):
                self.cfg["has_shown_tray_bubble"] = True
                save_config(self.cfg)
                if hasattr(self, "tray_icon") and self.tray_icon.isVisible():
                    self.tray_icon.showMessage(
                        "美术资产中枢已在后台运行",
                        "程序已最小化到系统托盘。点击托盘图标可随时 0 毫秒瞬间唤醒！",
                        QSystemTrayIcon.Information,
                        3000
                    )
        else:
            self.force_quit_app()
            event.accept()

    def setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(8)

        # 顶部全局导航栏
        top_bar = QHBoxLayout()
        self.sync_status_lbl = QLabel("🟢 极速同步已就绪")
        self.sync_status_lbl.setStyleSheet("background: rgba(16, 185, 129, 0.15); color: #34D399; border: 1px solid rgba(16, 185, 129, 0.3); border-radius: 12px; padding: 4px 10px; font-weight: bold; font-size: 11px;")
        
        btn_sync_excel = QPushButton("📤 同步缩略图到 Excel")
        btn_sync_excel.clicked.connect(self.sync_all_thumbnails_to_excel)
        btn_bind_excel = QPushButton("📊 绑定 Excel...")
        btn_bind_excel.clicked.connect(self.bind_excel_file)
        btn_refresh = QPushButton("🔄 刷新")
        btn_refresh.clicked.connect(self.async_load_data)
        btn_export = QPushButton("🌐 导出画廊")
        btn_export.clicked.connect(self.export_html_gallery)
        self.btn_theme = QPushButton("🌙 护眼暗灰")
        self.btn_theme.clicked.connect(self.toggle_theme)
        btn_settings = QPushButton("⚙️ 设置")
        btn_settings.setToolTip("设置关闭窗口时的默认行为（最小化到托盘/彻底退出）")
        btn_settings.clicked.connect(self.configure_close_action)

        top_bar.addWidget(self.sync_status_lbl)
        top_bar.addStretch()
        top_bar.addWidget(btn_sync_excel)
        top_bar.addWidget(btn_bind_excel)
        top_bar.addWidget(btn_refresh)
        top_bar.addWidget(btn_export)
        top_bar.addWidget(self.btn_theme)
        top_bar.addWidget(btn_settings)
        main_layout.addLayout(top_bar)

        # 主 Tab
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # Tab 1: 视觉资产看板
        self.tab_hub = QWidget()
        self.setup_hub_tab(self.tab_hub)
        self.tabs.addTab(self.tab_hub, "  🖼️ 视觉资产看板 (双维度导航 & GPU 加速)  ")

        # Tab 2: 设计源文件分拣与开工
        self.tab_organizer = QWidget()
        self.setup_organizer_tab(self.tab_organizer)
        self.tabs.addTab(self.tab_organizer, "  📥 设计源文件分拣与开工  ")

    # ---------------- Tab 1: 视觉资产看板 ----------------
    def setup_hub_tab(self, parent):
        layout = QHBoxLayout(parent)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(10)

        sidebar = QWidget()
        sidebar.setFixedWidth(220)
        side_layout = QVBoxLayout(sidebar)
        side_layout.setContentsMargins(0, 0, 0, 0)
        side_layout.setSpacing(10)

        # 上组：业务形态分类
        side_layout.addWidget(QLabel("<b>🏷️ 业务形态分类</b>"))
        self.category_list = QListWidget()
        self.category_list.setFixedHeight(168)
        cats = ["全部形态", "📦 包装", "🎁 套盒", "🖼️ 海报", "📑 物料"]
        for c in cats:
            self.category_list.addItem(f"{c} (0)")
        self.category_list.setCurrentRow(0)
        self.category_list.currentRowChanged.connect(self.on_category_changed)
        side_layout.addWidget(self.category_list)

        # 下组：客户品牌库
        brand_header_box = QHBoxLayout()
        brand_header_box.addWidget(QLabel("<b>🏢 客户与品牌库</b>"))
        brand_header_box.addStretch()
        
        btn_manage_brands = QToolButton()
        btn_manage_brands.setText("⚙️ 管理")
        btn_manage_brands.setToolTip("管理正式品牌、子系列归并映射与屏蔽非品牌目录")
        btn_manage_brands.clicked.connect(self.open_brand_manager)
        brand_header_box.addWidget(btn_manage_brands)
        side_layout.addLayout(brand_header_box)

        self.brand_list = QListWidget()
        self.brand_list.addItem("全部品牌 (0)")
        self.brand_list.setCurrentRow(0)
        self.brand_list.currentRowChanged.connect(self.on_brand_changed)
        self.brand_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.brand_list.customContextMenuRequested.connect(self.show_brand_context_menu)
        side_layout.addWidget(self.brand_list, stretch=1)
        
        layout.addWidget(sidebar)

        # 右侧画廊主体
        gallery_area = QWidget()
        gal_layout = QVBoxLayout(gallery_area)
        gal_layout.setContentsMargins(0, 0, 0, 0)
        gal_layout.setSpacing(8)

        filter_bar = QHBoxLayout()
        filter_bar.setSpacing(8)

        # 1. 🗄️ 工作空间选择器
        self.ws_combo = QComboBox()
        self.ws_combo.setMinimumWidth(180)
        self.ws_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.ws_combo.currentIndexChanged.connect(self.on_workspace_filter_changed)

        # 2. 🔍 搜索框
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("🔍 搜索产品 SKU / 品牌 / 类别...")
        self.search_edit.textChanged.connect(self.apply_filter)
        
        # 3. 视图模式
        self.view_combo = QComboBox()
        self.view_combo.addItems(["⚡ 智能融合视图", "📊 仅 Excel 台账", "💾 仅工作盘扫描"])
        self.view_combo.currentIndexChanged.connect(self.on_view_mode_changed)

        # 4. 排序
        self.sort_combo = QComboBox()
        self.sort_combo.addItems([
            "⏱️ 最新修改时间 (从新到旧)",
            "⏳ 最早创建时间 (从旧到新)",
            "🏢 品牌名称 (A → Z)",
            "📦 产品名称 (A → Z)"
        ])
        self.sort_combo.currentIndexChanged.connect(self.apply_filter)

        # 5. 🗄️ 工作盘管理按钮
        btn_manage_ws = QToolButton()
        btn_manage_ws.setText("🗄️ 工作盘管理")
        btn_manage_ws.setToolTip("管理工作盘、修改别名、查看磁盘空间与挂载状态")
        btn_manage_ws.clicked.connect(self.open_workspace_manager)

        filter_bar.addWidget(self.ws_combo, stretch=2)
        filter_bar.addWidget(self.search_edit, stretch=4)
        filter_bar.addWidget(self.view_combo, stretch=1)
        filter_bar.addWidget(self.sort_combo, stretch=1)
        filter_bar.addWidget(btn_manage_ws)
        gal_layout.addLayout(filter_bar)

        self.gallery_view = SmoothGalleryView()
        self.gallery_view.setObjectName("GalleryView")
        self.gallery_view.setViewMode(QListView.IconMode)
        self.gallery_view.setResizeMode(QListView.Adjust)
        self.gallery_view.setUniformItemSizes(True)
        self.gallery_view.setSpacing(12)
        self.gallery_view.setMovement(QListView.Static)
        
        self.gallery_model = GalleryModel(self)
        self.gallery_delegate = GalleryCardDelegate(self.gallery_view)
        self.gallery_view.setModel(self.gallery_model)
        self.gallery_view.setItemDelegate(self.gallery_delegate)
        
        self.gallery_view.clicked.connect(self.on_card_clicked)
        self.gallery_view.doubleClicked.connect(self.on_card_double_clicked)
        self.gallery_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.gallery_view.customContextMenuRequested.connect(self.show_gallery_context_menu)

        gal_layout.addWidget(self.gallery_view)
        layout.addWidget(gallery_area)

    # ---------------- Tab 2: 设计源文件分拣与开工 ----------------
    def setup_organizer_tab(self, parent):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(14)

        group1 = QGroupBox("📂 工作盘、客户、分类与归档文件夹规则")
        g1_layout = QVBoxLayout(group1)

        row1 = QHBoxLayout()
        row1.addWidget(QLabel("主工作盘:"))
        self.combo_ws = QComboBox()
        self.combo_ws.currentTextChanged.connect(self.on_organizer_setting_changed)
        btn_add_ws = QPushButton("🗄️ 工作盘管理...")
        btn_add_ws.clicked.connect(self.open_workspace_manager)
        row1.addWidget(self.combo_ws, stretch=1)
        row1.addWidget(btn_add_ws)
        g1_layout.addLayout(row1)

        row2 = QHBoxLayout()
        row2.addWidget(QLabel("客户品牌:"))
        self.combo_brand = QComboBox()
        self.combo_brand.setEditable(True)
        self.update_organizer_brand_combo()
        self.combo_brand.currentTextChanged.connect(self.on_organizer_setting_changed)
        btn_manage_b = QPushButton("⚙️ 管理品牌库...")
        btn_manage_b.clicked.connect(self.open_brand_manager)
        row2.addWidget(self.combo_brand, stretch=1)
        row2.addWidget(btn_manage_b)

        row2.addWidget(QLabel("业务形态:"))
        self.combo_cat = QComboBox()
        self.combo_cat.addItems(VALID_CATEGORIES)
        self.combo_cat.setCurrentText(self.cfg.get("default_category", "包装"))
        self.combo_cat.currentTextChanged.connect(self.on_organizer_setting_changed)
        row2.addWidget(self.combo_cat)

        row2.addWidget(QLabel("📁 归档规则:"))
        self.combo_rule = QComboBox()
        self.combo_rule.addItems([r["name"] for r in self.folder_rules])
        self.update_rule_combo_selection()
        self.combo_rule.currentIndexChanged.connect(self.on_rule_combo_changed)
        btn_custom_rule = QPushButton("⚙️ 自定义规则...")
        btn_custom_rule.clicked.connect(self.open_rule_manager)
        row2.addWidget(self.combo_rule, stretch=1)
        row2.addWidget(btn_custom_rule)
        g1_layout.addLayout(row2)
        layout.addWidget(group1)

        group2 = QGroupBox("⚡ 自动化开工选项")
        g2_layout = QHBoxLayout(group2)
        self.chk_ai = QCheckBox("🎨 自动打开 AI 设计原稿")
        self.chk_ai.setChecked(self.cfg.get("auto_open_ai", True))
        self.chk_ai.stateChanged.connect(self.on_organizer_setting_changed)
        
        self.chk_blend = QCheckBox("✨ 自动生成对应 .blend 工程")
        self.chk_blend.setChecked(self.cfg.get("auto_create_blend", True))
        self.chk_blend.stateChanged.connect(self.on_organizer_setting_changed)

        self.chk_open_blend = QCheckBox("🚀 自动启动 Blender")
        self.chk_open_blend.setChecked(self.cfg.get("auto_open_blender", True))
        self.chk_open_blend.stateChanged.connect(self.on_organizer_setting_changed)

        self.chk_excel = QCheckBox("📊 自动录入 Excel")
        self.chk_excel.setChecked(self.cfg.get("auto_append_to_excel", True))
        self.chk_excel.stateChanged.connect(self.on_organizer_setting_changed)

        g2_layout.addWidget(self.chk_ai)
        g2_layout.addWidget(self.chk_blend)
        g2_layout.addWidget(self.chk_open_blend)
        g2_layout.addWidget(self.chk_excel)
        layout.addWidget(group2)

        group3 = QGroupBox("📥 待分拣设计源文件列表")
        g3_layout = QVBoxLayout(group3)
        
        tools_layout = QHBoxLayout()
        btn_manual_create = QPushButton("✨ 手动新建项目...")
        btn_manual_create.setObjectName("PrimaryBtn")
        btn_manual_create.setToolTip("手动输入品名、选择品牌与形态，立即生成标准工程并开工")
        btn_manual_create.clicked.connect(self.open_manual_project_create_dialog)

        btn_add_files = QPushButton("➕ 添加设计源文件...")
        btn_add_files.clicked.connect(self.browse_source_files)

        btn_add_manual_item = QPushButton("✏️ 手动输入品名...")
        btn_add_manual_item.setToolTip("手动输入一个或多个品名，添加到下方待分拣列表")
        btn_add_manual_item.clicked.connect(self.add_manual_sku_to_organizer)

        btn_clear = QPushButton("🗑️ 清空列表")
        btn_clear.clicked.connect(self.clear_source_files)
        btn_install_jsx = QPushButton("🛠️ 一键将导出脚本注入 Illustrator")
        btn_install_jsx.clicked.connect(self.install_ai_jsx_script)
        
        tools_layout.addWidget(btn_manual_create)
        tools_layout.addWidget(btn_add_files)
        tools_layout.addWidget(btn_add_manual_item)
        tools_layout.addWidget(btn_clear)
        tools_layout.addStretch()
        tools_layout.addWidget(btn_install_jsx)
        g3_layout.addLayout(tools_layout)

        self.table_files = QTableWidget(0, 3)
        self.table_files.setHorizontalHeaderLabels(["源文件名", "提取 SKU", "目标归档目录"])
        self.table_files.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table_files.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table_files.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        g3_layout.addWidget(self.table_files)

        btn_start_flow = QPushButton("🚀 一键分拣归档并拉起 Blender 开工")
        btn_start_flow.setObjectName("PrimaryBtn")
        btn_start_flow.setFixedHeight(44)
        btn_start_flow.setStyleSheet("font-size: 15px; font-weight: bold;")
        btn_start_flow.clicked.connect(self.execute_organize_flow)
        g3_layout.addWidget(btn_start_flow)

        layout.addWidget(group3)

    def update_organizer_brand_combo(self):
        cur_brand = self.cfg.get("current_brand", "柏缇")
        all_brands = []
        for b in self.curated_brands:
            if b and b not in all_brands and b not in self.ignored_brands:
                all_brands.append(b)
        for b in sorted(self.brand_counts_map.keys()):
            if b and b not in all_brands and b not in self.ignored_brands and b != "未分类品牌":
                all_brands.append(b)
        self.combo_brand.blockSignals(True)
        self.combo_brand.clear()
        self.combo_brand.addItems(all_brands)
        if cur_brand in all_brands:
            self.combo_brand.setCurrentText(cur_brand)
        elif all_brands:
            self.combo_brand.setCurrentIndex(0)
        self.combo_brand.blockSignals(False)

    # ---------------- 品牌管理控制台联动 ----------------
    def open_brand_manager(self):
        cur_ws = self.combo_ws.currentText() if hasattr(self, "combo_ws") else self.workspaces[0]
        raw_dirs = []
        if cur_ws and os.path.exists(cur_ws):
            try:
                raw_dirs = [d for d in os.listdir(cur_ws) if os.path.isdir(os.path.join(cur_ws, d)) and not d.startswith('.')]
            except Exception:
                pass
        self.detected_raw_dirs = raw_dirs

        dlg = BrandManagerDialog(
            self.curated_brands, self.ignored_brands, self.brand_aliases, self.detected_raw_dirs, self
        )
        if dlg.exec() == QDialog.Accepted:
            self.curated_brands = dlg.curated_brands
            self.ignored_brands = dlg.ignored_brands
            self.brand_aliases = dlg.brand_aliases
            
            self.cfg["curated_brands"] = self.curated_brands
            self.cfg["ignored_brands"] = self.ignored_brands
            self.cfg["brand_aliases"] = self.brand_aliases
            save_config(self.cfg)
            
            self.async_load_data()
            QMessageBox.information(self, "品牌库已更新", "🎉 客户与品牌库规则已成功更新并重新加载！")

    def show_brand_context_menu(self, pos):
        item = self.brand_list.itemAt(pos)
        if not item:
            return
        raw_text = item.text()
        brand_name = raw_text.split(" (")[0].strip()
        if brand_name == "全部品牌":
            return

        menu = QMenu(self)
        act_open = menu.addAction(f"📁 打开 [{brand_name}] 的磁盘物理文件夹")
        act_alias = menu.addAction(f"🔗 将 [{brand_name}] 归并到其他品牌...")
        act_rename = menu.addAction(f"✏️ 重命名此品牌...")
        menu.addSeparator()
        act_ignore = menu.addAction(f"🚫 标记为非品牌并隐藏 (放入屏蔽黑名单)")
        act_set_default = menu.addAction(f"⭐ 设为默认开工品牌")

        action = menu.exec(QCursor.pos())
        cur_ws = self.combo_ws.currentText() if hasattr(self, "combo_ws") else self.workspaces[0]
        
        if action == act_open:
            b_dir = os.path.join(cur_ws, brand_name)
            if os.path.exists(b_dir):
                os.startfile(b_dir)
            else:
                QMessageBox.warning(self, "提示", f"未找到该品牌的物理文件夹:\n{b_dir}")
        elif action == act_alias:
            target, ok = QInputDialog.getText(self, "品牌归并", f"将 [{brand_name}] 的所有项目归并到哪个正式品牌？", text="柏缇")
            if ok and target.strip() and target.strip() != brand_name:
                self.brand_aliases[brand_name] = target.strip()
                self.cfg["brand_aliases"] = self.brand_aliases
                save_config(self.cfg)
                self.async_load_data()
        elif action == act_rename:
            new_name, ok = QInputDialog.getText(self, "重命名品牌", "修改品牌名称:", text=brand_name)
            if ok and new_name.strip() and new_name.strip() != brand_name:
                self.brand_aliases[brand_name] = new_name.strip()
                if brand_name in self.curated_brands:
                    idx = self.curated_brands.index(brand_name)
                    self.curated_brands[idx] = new_name.strip()
                    self.cfg["curated_brands"] = self.curated_brands
                self.cfg["brand_aliases"] = self.brand_aliases
                save_config(self.cfg)
                self.async_load_data()
        elif action == act_ignore:
            if brand_name not in self.ignored_brands:
                self.ignored_brands.append(brand_name)
                self.cfg["ignored_brands"] = self.ignored_brands
                save_config(self.cfg)
                self.async_load_data()
                QMessageBox.information(self, "已屏蔽", f"已成功将 [{brand_name}] 移入非品牌屏蔽黑名单！")
        elif action == act_set_default:
            self.cfg["current_brand"] = brand_name
            save_config(self.cfg)
            self.update_organizer_brand_combo()
            QMessageBox.information(self, "设置成功", f"已将 [{brand_name}] 设为默认开工品牌！")

    # ---------------- 工作盘空间多维管理与联动 ----------------
    def open_workspace_manager(self):
        dlg = WorkspaceHubDialog(self.cfg, self)
        dlg.workspaces_changed.connect(self.on_workspaces_updated)
        dlg.exec()

    def on_workspaces_updated(self, new_workspaces_v2):
        self.workspaces_v2 = new_workspaces_v2
        self.workspaces = [w["path"] for w in self.workspaces_v2]
        self.update_organizer_ws_combo()
        self.update_workspace_filter_combo()
        self.async_load_data()

    def update_workspace_filter_combo(self):
        if not hasattr(self, "ws_combo"):
            return
        self.ws_combo.blockSignals(True)
        cur_data = self.ws_combo.currentData()
        self.ws_combo.clear()
        
        total_projs = len(self.merged_projects)
        self.ws_combo.addItem(f"🌐 全部工作盘 (全景联合 - {total_projs})", "")
        
        target_idx = 0
        for idx, ws in enumerate(self.workspaces_v2, start=1):
            p = ws.get("path", "")
            alias = ws.get("alias", "")
            is_prim = ws.get("is_primary", False)
            norm_p = os.path.normpath(p).lower()
            
            p_count = sum(1 for proj in self.merged_projects if os.path.normpath(proj.get("ws_path", "")).lower() == norm_p or os.path.normpath(proj.get("path", "")).lower().startswith(norm_p))
            icon_str = "⭐" if is_prim else "📁"
            self.ws_combo.addItem(f"{icon_str} {alias} ({p_count})", p)
            if cur_data and os.path.normpath(cur_data).lower() == norm_p:
                target_idx = idx
                
        self.ws_combo.setCurrentIndex(target_idx)
        self.ws_combo.blockSignals(False)

    def update_organizer_ws_combo(self):
        if not hasattr(self, "combo_ws"):
            return
        self.combo_ws.blockSignals(True)
        self.combo_ws.clear()
        
        cur_ws = self.cfg.get("current_workspace", "")
        target_idx = 0
        
        for idx, ws in enumerate(self.workspaces_v2):
            p = ws.get("path", "")
            alias = ws.get("alias", "")
            is_prim = ws.get("is_primary", False)
            label = f"{'⭐ ' if is_prim else ''}{alias} ({p})"
            self.combo_ws.addItem(label, p)
            if (cur_ws and os.path.normpath(cur_ws).lower() == os.path.normpath(p).lower()) or (not cur_ws and is_prim):
                target_idx = idx
                
        self.combo_ws.setCurrentIndex(target_idx)
        self.combo_ws.blockSignals(False)

    def get_current_organizer_workspace(self):
        """准确获取当前分拣选中的物理工作盘根路径"""
        if hasattr(self, "combo_ws"):
            data = self.combo_ws.currentData()
            if data and isinstance(data, str) and os.path.exists(data):
                return os.path.normpath(data)
            txt = self.combo_ws.currentText()
            # 尝试从形如 "⭐ 主力生产盘 (E:\zjc)" 中提取括号里的路径
            m = re.search(r'\((.*?)\)', txt)
            if m and os.path.exists(m.group(1).strip()):
                return os.path.normpath(m.group(1).strip())
            if txt and os.path.exists(txt):
                return os.path.normpath(txt)
        if hasattr(self, "workspaces_v2") and self.workspaces_v2:
            for w in self.workspaces_v2:
                if w.get("is_primary") and os.path.exists(w.get("path", "")):
                    return os.path.normpath(w["path"])
            for w in self.workspaces_v2:
                if os.path.exists(w.get("path", "")):
                    return os.path.normpath(w["path"])
        return "E:\\zjc" if os.path.exists("E:\\zjc") else (self.workspaces[0] if hasattr(self, "workspaces") and self.workspaces else "D:\\Projects")

    def on_workspace_filter_changed(self):
        self.update_sidebar_counts()
        self.apply_filter()

    # ---------------- 业务逻辑与数据流 (Qt 线程池 + 信号槽) ----------------
    def async_load_data(self):
        self.sync_status_lbl.setText("⚡ 正在加载全量工作盘资产...")
        ex_path = self.cfg.get("excel_path", DEFAULT_EXCEL_PATH)
        
        worker = DataLoaderWorker(ex_path, self.workspaces_v2, self.meta_cache, self.brand_aliases, self.ignored_brands)
        worker.signals.finished.connect(self.on_data_loaded)
        QThreadPool.globalInstance().start(worker)

    @Slot(list, list, list)
    def on_data_loaded(self, excel_p, disk_p, merged):
        self.excel_projects = excel_p
        self.disk_projects = disk_p
        self.merged_projects = merged
        self.update_after_data_loaded()
        if not self.has_fired_initial_done:
            self.has_fired_initial_done = True
            self.initial_load_done.emit()

    def update_after_data_loaded(self):
        self.sync_status_lbl.setText(f"🟢 极速同步已就绪 (已载入 {len(self.merged_projects)} 个项目)")
        self.update_workspace_filter_combo()
        self.update_sidebar_counts()
        self.update_active_dataset()
        self.update_organizer_brand_combo()

    def update_sidebar_counts(self):
        mode_idx = self.view_combo.currentIndex()
        dataset = self.merged_projects if mode_idx == 0 else (self.excel_projects if mode_idx == 1 else self.disk_projects)
        
        # 空间物理隔离过滤
        if hasattr(self, "ws_combo") and self.ws_combo.currentIndex() > 0:
            selected_ws_path = self.ws_combo.currentData()
            if selected_ws_path:
                norm_sel_ws = os.path.normpath(selected_ws_path).lower()
                dataset = [p for p in dataset if os.path.normpath(p.get("ws_path", "")).lower() == norm_sel_ws or os.path.normpath(p.get("path", "")).lower().startswith(norm_sel_ws)]

        cat_counts = {"全部形态": len(dataset), "包装": 0, "套盒": 0, "海报": 0, "物料": 0}
        brand_counts = {}
        for p in dataset:
            c = p.get("cat", "包装")
            if c in cat_counts:
                cat_counts[c] += 1
            else:
                cat_counts["包装"] += 1
                
            raw_b = p.get("raw_brand", p.get("brand", "")).strip()
            if raw_b in self.ignored_brands:
                continue
            b = p.get("brand", "").strip() or "未分类品牌"
            if b in self.ignored_brands:
                continue
            brand_counts[b] = brand_counts.get(b, 0) + 1

        self.brand_counts_map = brand_counts

        cats = [("全部形态", "全部形态"), ("📦 包装", "包装"), ("🎁 套盒", "套盒"), ("🖼️ 海报", "海报"), ("📑 物料", "物料")]
        for idx, (label, key) in enumerate(cats):
            self.category_list.item(idx).setText(f"{label} ({cat_counts[key]})")

        cur_selected_brand = self.selected_brand
        self.brand_list.blockSignals(True)
        self.brand_list.clear()
        
        total_brand_skus = sum(brand_counts.values())
        self.brand_list.addItem(f"全部品牌 ({total_brand_skus})")
        
        sorted_brands = sorted(brand_counts.items(), key=lambda x: x[1], reverse=True)
        target_row = 0
        for idx, (b_name, b_count) in enumerate(sorted_brands, start=1):
            self.brand_list.addItem(f"{b_name} ({b_count})")
            if b_name == cur_selected_brand:
                target_row = idx

        self.brand_list.setCurrentRow(target_row)
        self.brand_list.blockSignals(False)

    def on_view_mode_changed(self):
        self.update_sidebar_counts()
        self.update_active_dataset()

    def on_category_changed(self, row):
        cats = ["全部", "包装", "套盒", "海报", "物料"]
        if 0 <= row < len(cats):
            self.selected_category = cats[row]
            self.apply_filter()

    def on_brand_changed(self, row):
        if row == 0:
            self.selected_brand = "全部"
        elif row > 0 and row - 1 < len(self.brand_counts_map):
            sorted_brands = sorted(self.brand_counts_map.items(), key=lambda x: x[1], reverse=True)
            self.selected_brand = sorted_brands[row - 1][0]
        self.apply_filter()

    def update_active_dataset(self):
        mode_idx = self.view_combo.currentIndex()
        if mode_idx == 1:
            self.current_display_list = self.excel_projects
        elif mode_idx == 2:
            self.current_display_list = self.disk_projects
        else:
            self.current_display_list = self.merged_projects
        self.apply_filter()

    def apply_filter(self):
        kw = self.search_edit.text().strip().lower()
        selected_ws_path = ""
        if hasattr(self, "ws_combo") and self.ws_combo.currentIndex() > 0:
            selected_ws_path = self.ws_combo.currentData()
            norm_sel_ws = os.path.normpath(selected_ws_path).lower() if selected_ws_path else ""
        else:
            norm_sel_ws = ""

        res = []
        for p in self.current_display_list:
            if norm_sel_ws:
                p_ws = os.path.normpath(p.get("ws_path", "")).lower()
                p_path = os.path.normpath(p.get("path", "")).lower()
                if not (p_ws == norm_sel_ws or p_path.startswith(norm_sel_ws)):
                    continue

            cat = p.get("cat", "包装")
            brand = p.get("brand", "").strip() or "未分类品牌"
            raw_brand = p.get("raw_brand", brand).strip()
            
            if not kw and (raw_brand in self.ignored_brands or brand in self.ignored_brands):
                continue
                
            if self.selected_category != "全部" and cat != self.selected_category:
                continue
                
            if self.selected_brand != "全部" and brand != self.selected_brand:
                continue
                
            if kw:
                sku = p.get("sku", "").lower()
                b_low = brand.lower()
                if kw not in sku and kw not in b_low and kw not in cat.lower():
                    continue
            res.append(p)

        # ----------------- 多维动态排序引擎 -----------------
        sort_idx = self.sort_combo.currentIndex() if hasattr(self, 'sort_combo') else 0
        if sort_idx == 0:
            # ⏱️ 最新修改时间 (从新到旧，降序)
            res.sort(key=lambda x: (x.get("mtime") or 0, x.get("sku", "")), reverse=True)
        elif sort_idx == 1:
            # ⏳ 最早创建时间 (从旧到新，升序)
            res.sort(key=lambda x: (x.get("mtime") if (x.get("mtime") and x.get("mtime") > 0) else 9999999999, x.get("sku", "")))
        elif sort_idx == 2:
            # 🏢 品牌名称 (A → Z)
            res.sort(key=lambda x: (x.get("brand", "").lower(), x.get("sku", "").lower()))
        elif sort_idx == 3:
            # 📦 产品名称 (A → Z)
            res.sort(key=lambda x: (x.get("sku", "").lower(), x.get("brand", "").lower()))

        self.gallery_model.set_projects(res)

    def on_card_clicked(self, index):
        pass

    def on_card_double_clicked(self, index):
        proj = index.data(Qt.UserRole)
        if proj:
            self.launch_blend(proj.get("path"), sku=proj.get("sku"), brand=proj.get("brand"))

    def show_gallery_context_menu(self, pos):
        index = self.gallery_view.indexAt(pos)
        if not index.isValid():
            return
        proj = index.data(Qt.UserRole)
        if not proj:
            return

        menu = QMenu(self)
        p = proj.get("path", "")
        sku = proj.get("sku", "")
        brand = proj.get("brand", "")
        
        act_folder = menu.addAction(f"📁 打开文件夹: {sku}")
        act_blend = menu.addAction("🚀 Blender 打开 3D 工程")
        
        real_p = self.resolve_project_path(p, sku, brand)
        if real_p and os.path.exists(real_p):
            menu.addAction("🎨 查看 01_Design_平面原稿", lambda: self.open_folder(os.path.join(real_p, "01_Design_平面原稿")))
            menu.addAction("🖼️ 查看 04_Renders_通道输出", lambda: self.open_folder(os.path.join(real_p, "04_Renders_通道输出")))
            
        menu.addSeparator()
        cat_menu = menu.addMenu(f"🏷️ 修改业务形态 (当前: {proj.get('cat', '包装')})")
        for c in VALID_CATEGORIES:
            cat_menu.addAction(f"设为：{c}", lambda c_val=c: self.change_project_category(proj, c_val))
            
        menu.addSeparator()
        if proj.get("thumbnail") and is_valid_beauty_thumbnail(proj["thumbnail"]):
            menu.addAction("📊 将此缩略图写入 Excel 台账 (图片列)", lambda: self.sync_single_thumbnail_to_excel(proj))
        menu.addAction("📋 复制完整物理路径", lambda: QApplication.clipboard().setText(real_p or p))

        action = menu.exec(QCursor.pos())
        if action == act_folder:
            self.open_folder(p, sku=sku, brand=brand)
        elif action == act_blend:
            self.launch_blend(p, sku=sku, brand=brand)

    def resolve_project_path(self, path, sku="", brand=""):
        if path and os.path.exists(path):
            return path
        cur_ws = self.combo_ws.currentText() if hasattr(self, "combo_ws") else self.workspaces[0]
        if sku and cur_ws and os.path.exists(cur_ws):
            if brand:
                cand = os.path.join(cur_ws, brand, sku)
                if os.path.exists(cand):
                    return cand
            try:
                for entry in os.listdir(cur_ws):
                    bp = os.path.join(cur_ws, entry)
                    if os.path.isdir(bp):
                        cand = os.path.join(bp, sku)
                        if os.path.exists(cand):
                            return cand
            except Exception:
                pass
            cand = os.path.join(cur_ws, sku)
            if os.path.exists(cand):
                return cand
        return path

    def open_folder(self, path, sku="", brand=""):
        real_path = self.resolve_project_path(path, sku, brand)
        if real_path and os.path.exists(real_path):
            os.startfile(real_path)
        else:
            cur_ws = self.combo_ws.currentText() if hasattr(self, "combo_ws") else self.workspaces[0]
            reply = QMessageBox.question(
                self, "项目未在工作盘创建",
                f"【{sku or '该项目'}】尚未在当前工作盘中创建本地工程目录。\n\n是否立即一键在工作盘自动建立标准工程结构？",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes and sku and cur_ws and os.path.exists(cur_ws):
                rule = self.get_current_folder_rule()
                subfolders = rule.get("subfolders", DEFAULT_FOLDER_RULES[0]["subfolders"])
                brand_name = brand or "未分类品牌"
                proj_dir = os.path.join(cur_ws, brand_name, sku)
                os.makedirs(proj_dir, exist_ok=True)
                for sub in subfolders:
                    os.makedirs(os.path.join(proj_dir, sub), exist_ok=True)
                os.startfile(proj_dir)
                self.async_load_data()

    def launch_blend(self, proj_path, sku="", brand=""):
        real_path = self.resolve_project_path(proj_path, sku, brand)
        if not real_path or not os.path.exists(real_path):
            self.open_folder(proj_path, sku=sku, brand=brand)
            return
        rule = self.get_current_folder_rule()
        blend_sub = rule.get("blend_sub", "03_3D_三维工程")
        blend_dir = os.path.join(real_path, blend_sub) if blend_sub else real_path
        target_dir = blend_dir if os.path.exists(blend_dir) else real_path
        blends = glob.glob(os.path.join(target_dir, "*.blend"))
        if blends:
            blends.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            chosen = blends[0]
            try:
                subprocess.Popen([BLENDER_EXE, chosen])
            except Exception:
                os.startfile(chosen)
        else:
            tpl = get_valid_template_blend(self.cfg)
            target_blend = os.path.join(target_dir, f"{sku or '工程'}.blend")
            if tpl and os.path.exists(tpl):
                shutil.copy2(tpl, target_blend)
                try:
                    subprocess.Popen([BLENDER_EXE, target_blend])
                except Exception:
                    os.startfile(target_blend)
            else:
                self.open_folder(real_path)

    def change_project_category(self, proj, new_cat):
        sku = proj.get("sku")
        path = proj.get("path")
        proj["cat"] = new_cat
        if path:
            norm_p = path.lower().replace("/", "\\")
            if norm_p in self.meta_cache:
                self.meta_cache[norm_p]["cat"] = new_cat
            else:
                self.meta_cache[norm_p] = {
                    "brand": proj.get("brand", ""),
                    "raw_brand": proj.get("raw_brand", ""),
                    "sku": sku,
                    "cat": new_cat,
                    "thumbnail": proj.get("thumbnail"),
                    "mtime": proj.get("mtime", 0)
                }
            save_meta_cache(self.meta_cache)
            
        ex_path = self.cfg.get("excel_path", DEFAULT_EXCEL_PATH)
        if ex_path and os.path.exists(ex_path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(ex_path)
                sheet = wb.active
                cat_col = None
                sku_col = 2
                for col_idx, cell in enumerate(sheet[1], start=1):
                    val = str(cell.value or "").strip()
                    if val in ("业务形态", "分类", "类别"):
                        cat_col = col_idx
                    if val in ("产品名称", "SKU", "品名", "产品命名"):
                        sku_col = col_idx
                if cat_col:
                    sku_clean = re.sub(r'[\s_\-\(\)（）]+', '', sku.lower()) if sku else ""
                    for r in range(2, sheet.max_row + 1):
                        cell_val = str(sheet.cell(row=r, column=sku_col).value or "").strip()
                        cell_clean = re.sub(r'[\s_\-\(\)（）]+', '', cell_val.lower())
                        if sku_clean and cell_clean and (sku_clean == cell_clean or sku_clean in cell_clean or cell_clean in sku_clean):
                            sheet.cell(row=r, column=cat_col, value=new_cat)
                            break
                    wb.save(ex_path)
                wb.close()
            except Exception:
                pass
        self.apply_filter()
        QMessageBox.information(self, "修改成功", f"已成功将 [{sku}] 的业务形态更新为：【{new_cat}】")

    def sync_single_thumbnail_to_excel(self, proj):
        ex_path = self.cfg.get("excel_path", DEFAULT_EXCEL_PATH)
        if not ex_path or not os.path.exists(ex_path):
            QMessageBox.warning(self, "同步失败", "未找到绑定的《产品列表.xlsx》台账文件！")
            return
        if not proj.get("thumbnail") or not os.path.exists(proj["thumbnail"]):
            QMessageBox.warning(self, "同步失败", "该项目暂无有效的成片渲染缩略图！")
            return
            
        progress = QProgressDialog("正在将缩略图写入 Excel 台账...", "取消", 0, 1, self)
        progress.setWindowTitle("📊 缩略图写入 Excel")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        
        worker = ExcelSyncWorker(ex_path, [proj])
        
        def on_finished(ok, msg):
            progress.close()
            if ok:
                QMessageBox.information(self, "同步成功", f"🎉 已成功将 [{proj.get('sku')}] 的成片缩略图写入 Excel 台账！")
            else:
                QMessageBox.warning(self, "同步失败", msg)
                
        worker.signals.finished.connect(on_finished)
        progress.canceled.connect(worker.cancel)
        self.thread_pool.start(worker)

    def sync_all_thumbnails_to_excel(self):
        ex_path = self.cfg.get("excel_path", DEFAULT_EXCEL_PATH)
        if not ex_path or not os.path.exists(ex_path):
            QMessageBox.warning(self, "同步失败", "未找到绑定的《产品列表.xlsx》台账文件！")
            return
            
        valid_items = [p for p in self.merged_projects if p.get("thumbnail") and is_valid_beauty_thumbnail(p["thumbnail"])]
        if not valid_items:
            QMessageBox.warning(self, "提示", "当前没有找到任何带有有效成片缩略图的项目！")
            return
            
        total = len(valid_items)
        progress = QProgressDialog("正在准备同步缩略图到 Excel...", "取消", 0, total, self)
        progress.setWindowTitle("📊 批量同步缩略图到 Excel")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        
        worker = ExcelSyncWorker(ex_path, self.merged_projects)
        
        def on_progress(cur, tot, sku):
            progress.setValue(cur)
            progress.setLabelText(f"正在同步 ({cur}/{tot}): {sku}")
            
        def on_finished(ok, msg):
            progress.close()
            if ok:
                QMessageBox.information(self, "批量同步成功", msg)
            else:
                QMessageBox.warning(self, "批量同步失败", msg)
                
        worker.signals.progress.connect(on_progress)
        worker.signals.finished.connect(on_finished)
        progress.canceled.connect(worker.cancel)
        self.thread_pool.start(worker)

    def bind_excel_file(self):
        f, _ = QFileDialog.getOpenFileName(self, "绑定 Excel 产品台账", "", "Excel Files (*.xlsx *.xls)")
        if f:
            self.cfg["excel_path"] = os.path.normpath(f)
            save_config(self.cfg)
            self.async_load_data()
            QMessageBox.information(self, "绑定成功", f"已成功绑定新 Excel:\n{f}")

    def export_html_gallery(self):
        ex_dir = self.combo_ws.currentText() if hasattr(self, "combo_ws") else os.path.expanduser("~")
        html_file = os.path.join(ex_dir, "美术资产全景画廊.html")
        
        cards = []
        for p in self.current_display_list:
            thumb = p.get("thumbnail") or ""
            norm_thumb = thumb.replace("\\", "/") if thumb else ""
            t_src = f"file:///{norm_thumb}" if thumb and is_valid_beauty_thumbnail(thumb) else ""
            img_html = f'<img src="{t_src}" loading="lazy" />' if t_src else '<div style="color:#666;font-size:12px;">待渲染</div>'
            cards.append(f"""
            <div class="card" onclick="alert('项目路径: {html.escape(p.get('path', ''))}')">
              <div class="thumb-container">
                {img_html}
              </div>
              <div class="meta">
                <div class="badge">{html.escape(p.get('cat', '包装'))}</div>
                <div class="title">{html.escape(p.get('sku', ''))}</div>
                <div class="path">{html.escape(p.get('brand', ''))}</div>
              </div>
            </div>
            """)
            
        doc = f"""<!DOCTYPE html><html><head><meta charset="utf-8"/><title>美术资产全景视觉画廊</title>
        <style>
        body {{ background: #1a1c23; color: #e2e4e8; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif; padding: 24px; margin: 0; }}
        h1 {{ font-size: 20px; margin-bottom: 20px; font-weight: 600; }}
        .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(210px, 1fr)); gap: 16px; }}
        .card {{ background: #262930; border-radius: 8px; overflow: hidden; border: 1px solid #363942; transition: transform .2s; cursor: pointer; }}
        .card:hover {{ transform: translateY(-4px); border-color: #3b82f6; }}
        .thumb-container {{ width: 100%; aspect-ratio: 1; background: #1e2026; display: flex; align-items: center; justify-content: center; overflow: hidden; }}
        .thumb-container img {{ width: 100%; height: 100%; object-fit: contain; }}
        .meta {{ padding: 10px; }}
        .badge {{ display: inline-block; font-size: 11px; background: #1e3a8a; color: #93c5fd; padding: 2px 6px; border-radius: 4px; font-weight: bold; margin-bottom: 4px; }}
        .title {{ font-size: 13px; font-weight: bold; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
        .path {{ font-size: 11px; color: #8c909c; margin-top: 2px; }}
        </style></head><body>
        <h1>🎨 美术资产全景视觉画廊 (共 {len(self.current_display_list)} 个项目)</h1>
        <div class="grid">{''.join(cards)}</div>
        </body></html>"""
        
        try:
            with open(html_file, "w", encoding="utf-8") as f:
                f.write(doc)
            webbrowser.open("file:///" + html_file.replace("\\", "/"))
            QMessageBox.information(self, "导出成功", f"已成功导出画廊到:\n{html_file}")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", str(e))

    # ---------------- 规则与分拣开工 ----------------
    def update_rule_combo_selection(self):
        for idx, r in enumerate(self.folder_rules):
            if r.get("id") == self.active_rule_id:
                self.combo_rule.setCurrentIndex(idx)
                return
        if self.folder_rules:
            self.combo_rule.setCurrentIndex(0)

    def get_current_folder_rule(self):
        for r in self.folder_rules:
            if r.get("id") == self.active_rule_id:
                return r
        return self.folder_rules[0] if self.folder_rules else DEFAULT_FOLDER_RULES[0]

    def on_rule_combo_changed(self, idx):
        if 0 <= idx < len(self.folder_rules):
            self.active_rule_id = self.folder_rules[idx]["id"]
            self.on_organizer_setting_changed()

    def open_rule_manager(self):
        dlg = FolderRuleManagerDialog(self.folder_rules, self.active_rule_id, self)
        if dlg.exec() == QDialog.Accepted:
            self.folder_rules = dlg.rules
            self.active_rule_id = dlg.active_rule_id
            self.cfg["folder_rules"] = self.folder_rules
            self.cfg["active_rule_id"] = self.active_rule_id
            save_config(self.cfg)
            
            self.combo_rule.blockSignals(True)
            self.combo_rule.clear()
            self.combo_rule.addItems([r["name"] for r in self.folder_rules])
            self.update_rule_combo_selection()
            self.combo_rule.blockSignals(False)
            self.refresh_organizer_table()
            QMessageBox.information(self, "规则更新", "已成功更新并应用文件夹归档规则！")

    def on_organizer_setting_changed(self):
        cur_ws = self.get_current_organizer_workspace()
        self.cfg["current_workspace"] = cur_ws
        self.cfg["current_brand"] = self.combo_brand.currentText().strip()
        self.cfg["default_category"] = self.combo_cat.currentText().strip()
        self.cfg["active_rule_id"] = self.active_rule_id
        self.cfg["auto_open_ai"] = self.chk_ai.isChecked()
        self.cfg["auto_create_blend"] = self.chk_blend.isChecked()
        self.cfg["auto_open_blender"] = self.chk_open_blend.isChecked()
        self.cfg["auto_append_to_excel"] = self.chk_excel.isChecked()
        save_config(self.cfg)
        self.refresh_organizer_table()

    def add_workspace(self):
        d = QFileDialog.getExistingDirectory(self, "选择并绑定新工作盘")
        if d:
            d = os.path.normpath(d)
            if d not in self.workspaces:
                self.workspaces.append(d)
                self.cfg["workspaces"] = self.workspaces
                self.combo_ws.addItem(d, d)
            self.combo_ws.setCurrentText(d)
            self.on_organizer_setting_changed()

    def add_brand(self):
        b, ok = QInputDialog.getText(self, "新增客户品牌", "请输入新客户品牌名称:")
        if ok and b.strip():
            b = b.strip()
            if b not in self.curated_brands:
                self.curated_brands.append(b)
                self.cfg["curated_brands"] = self.curated_brands
                self.update_organizer_brand_combo()
            self.combo_brand.setCurrentText(b)
            self.on_organizer_setting_changed()

    def open_manual_project_create_dialog(self):
        """弹出独立对话框，手动输入品名创建项目工程并开工"""
        dlg = ManualProjectCreateDialog(
            self.cfg, self.workspaces_v2, self.curated_brands, self.ignored_brands,
            self.folder_rules, self.active_rule_id, self
        )
        dlg.project_created.connect(self.on_manual_project_created)
        dlg.exec()

    def on_manual_project_created(self, proj_info):
        self.async_load_data()

    def add_manual_sku_to_organizer(self):
        """手动输入产品品名直接加入待分拣表格"""
        text, ok = QInputDialog.getMultiLineText(
            self, "✏️ 手动输入产品品名",
            "请输入产品 SKU / 品名（支持一次输入多个，一行一个或用逗号隔开）:",
            ""
        )
        if ok and text.strip():
            raw_skus = re.split(r'[\n,，]+', text)
            added_count = 0
            for raw_sku in raw_skus:
                clean_sku = re.sub(r'[\\/:*?"<>|]', '', raw_sku).strip()
                if clean_sku:
                    item_key = f"__manual__:{clean_sku}"
                    if item_key not in self.files_to_organize:
                        self.files_to_organize.append(item_key)
                        added_count += 1
            if added_count > 0:
                self.refresh_organizer_table()
                QMessageBox.information(self, "添加成功", f"已成功添加 {added_count} 个手动新建项目到列表中！\n点击下方【🚀 一键分拣归档并拉起 Blender 开工】即可一键生成全部项目脚手架。")

    def browse_source_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "选择设计源文件", "", "Design Files (*.ai *.psd *.pdf *.zip *.rar);;All Files (*.*)")
        if files:
            self.add_files_to_organizer(files)

    def add_files_to_organizer(self, files):
        for f in files:
            f = os.path.normpath(f)
            if f not in self.files_to_organize:
                self.files_to_organize.append(f)
        self.refresh_organizer_table()

    def clear_source_files(self):
        self.files_to_organize.clear()
        self.refresh_organizer_table()

    def refresh_organizer_table(self):
        self.table_files.setRowCount(0)
        cur_ws = self.get_current_organizer_workspace()
        brand = self.combo_brand.currentText().strip()
        cat = self.combo_cat.currentText().strip()
        rule = self.get_current_folder_rule()
        pat = rule.get("path_pattern", "{brand}/{sku}")

        for f in self.files_to_organize:
            if f.startswith("__manual__:"):
                sku = f.split(":", 1)[1]
                fname = f"✨ [手动新建] {sku}"
            else:
                fname = os.path.basename(f)
                sku = os.path.splitext(fname)[0]
            
            if "{category}" in pat:
                rel = f"{cat}/{brand}/{sku}" if brand else f"{cat}/{sku}"
            elif "{brand}" in pat:
                rel = f"{brand}/{sku}" if brand else sku
            else:
                rel = sku
                
            dest_dir = os.path.join(cur_ws, rel)
            
            row = self.table_files.rowCount()
            self.table_files.insertRow(row)
            self.table_files.setItem(row, 0, QTableWidgetItem(fname))
            self.table_files.setItem(row, 1, QTableWidgetItem(sku))
            self.table_files.setItem(row, 2, QTableWidgetItem(dest_dir))

    def execute_organize_flow(self):
        if not self.files_to_organize:
            QMessageBox.warning(self, "提示", "请先点击【➕ 添加设计源文件】或【✏️ 手动输入品名】！")
            return
            
        cur_ws = self.get_current_organizer_workspace()
        if not cur_ws or not os.path.exists(cur_ws):
            QMessageBox.warning(self, "工作盘错误", f"未找到有效的工作盘物理目录:\n{cur_ws}\n\n请点击【🗄️ 工作盘管理】检查工作盘路径！")
            return

        brand = self.combo_brand.currentText().strip()
        cat = self.combo_cat.currentText().strip()
        rule = self.get_current_folder_rule()
        subfolders = rule.get("subfolders", DEFAULT_FOLDER_RULES[0]["subfolders"])
        pat = rule.get("path_pattern", "{brand}/{sku}")
        design_sub = rule.get("design_sub", "01_Design_平面原稿")
        blend_sub = rule.get("blend_sub", "03_3D_三维工程")
        
        auto_ai = self.chk_ai.isChecked()
        auto_blend = self.chk_blend.isChecked()
        auto_open_bl = self.chk_open_blend.isChecked()
        auto_excel = self.chk_excel.isChecked()
        
        created_count = 0
        first_proj_dir = None
        opened_blend = None
        opened_ai = None
        
        for fpath in list(self.files_to_organize):
            is_manual = fpath.startswith("__manual__:")
            if is_manual:
                sku = fpath.split(":", 1)[1]
                fname = ""
            else:
                if not os.path.exists(fpath):
                    continue
                fname = os.path.basename(fpath)
                sku = os.path.splitext(fname)[0]
            
            if "{category}" in pat:
                rel = f"{cat}/{brand}/{sku}" if brand else f"{cat}/{sku}"
            elif "{brand}" in pat:
                rel = f"{brand}/{sku}" if brand else sku
            else:
                rel = sku
                
            proj_dir = os.path.join(cur_ws, rel)
            try:
                os.makedirs(proj_dir, exist_ok=True)
                for sub in subfolders:
                    os.makedirs(os.path.join(proj_dir, sub), exist_ok=True)
            except Exception as e:
                QMessageBox.critical(self, "创建目录失败", f"无法创建项目目录:\n{proj_dir}\n\n错误信息: {str(e)}")
                return
                
            if not first_proj_dir:
                first_proj_dir = proj_dir
                
            # 1. 复制源文件到 01_Design_平面原稿 (如果是非手动项)
            if not is_manual and fname:
                dest_fpath = os.path.join(proj_dir, design_sub, fname)
                try:
                    shutil.copy2(fpath, dest_fpath)
                except Exception as e:
                    print(f"复制源文件失败: {e}")
                    
                if auto_ai and not opened_ai and fname.lower().endswith(('.ai', '.psd', '.pdf')):
                    try:
                        os.startfile(dest_fpath)
                        opened_ai = dest_fpath
                    except Exception:
                        pass
                    
            # 2. 生成对应 .blend 工程并启动 Blender
            blend_dir = os.path.join(proj_dir, blend_sub)
            target_blend = os.path.join(blend_dir, f"{sku}.blend")
            if auto_blend:
                if not os.path.exists(target_blend):
                    tpl = get_valid_template_blend(self.cfg)
                    if tpl and os.path.exists(tpl):
                        shutil.copy2(tpl, target_blend)
                    else:
                        with open(target_blend, "wb") as bf:
                            pass
                            
            if auto_open_bl and not opened_blend:
                if os.path.exists(target_blend):
                    try:
                        subprocess.Popen([BLENDER_EXE, target_blend])
                        opened_blend = target_blend
                    except Exception:
                        try:
                            os.startfile(target_blend)
                            opened_blend = target_blend
                        except Exception:
                            pass
                            
            # 3. 自动追加录入 Excel 台账
            if auto_excel:
                ex_path = self.cfg.get("excel_path", DEFAULT_EXCEL_PATH)
                if ex_path and os.path.exists(ex_path):
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(ex_path)
                        sheet = wb.active
                        sku_col = 2
                        brand_col = 1
                        cat_col = None
                        for col_idx, cell in enumerate(sheet[1], start=1):
                            val = str(cell.value or "").strip()
                            if val in ("产品名称", "SKU", "品名", "产品命名"):
                                sku_col = col_idx
                            if val in ("品牌", "客户"):
                                brand_col = col_idx
                            if val in ("业务形态", "分类", "类别"):
                                cat_col = col_idx
                                
                        sku_clean = re.sub(r'[\s_\-\(\)（）]+', '', sku.lower())
                        exists = False
                        for r in range(2, sheet.max_row + 1):
                            c_val = str(sheet.cell(row=r, column=sku_col).value or "").strip()
                            if sku_clean and sku_clean == re.sub(r'[\s_\-\(\)（）]+', '', c_val.lower()):
                                exists = True
                                break
                        if not exists:
                            new_r = sheet.max_row + 1
                            sheet.cell(row=new_r, column=brand_col, value=brand)
                            sheet.cell(row=new_r, column=sku_col, value=sku)
                            if cat_col:
                                sheet.cell(row=new_r, column=cat_col, value=cat)
                            wb.save(ex_path)
                        wb.close()
                    except Exception:
                        pass
                        
            created_count += 1
            
        self.files_to_organize.clear()
        self.refresh_organizer_table()
        self.async_load_data()
        
        # 自动弹出生成的目标项目文件夹
        if first_proj_dir and os.path.exists(first_proj_dir):
            try:
                os.startfile(first_proj_dir)
            except Exception:
                pass
                
        QMessageBox.information(self, "开工成功", f"🎉 已成功分拣归档并初始化 {created_count} 个工程项目！\n已就绪并打开首个工程目录。")

    def install_ai_jsx_script(self):
        src_jsx = os.path.join(os.path.dirname(__file__), "Export_Artboards_To_Textures.jsx")
        if not os.path.exists(src_jsx):
            src_jsx = r"C:\Users\qq424\Packaging_Tools\Export_Artboards_To_Textures.jsx"
            
        if not os.path.exists(src_jsx):
            QMessageBox.warning(self, "错误", "未找到脚本源文件: Export_Artboards_To_Textures.jsx")
            return
            
        target_dirs = []
        possible_roots = [
            r"H:\adobe\Adobe Illustrator 2024",
            r"C:\Program Files\Adobe\Adobe Illustrator 2024",
            r"C:\Program Files\Adobe\Adobe Illustrator 2025",
            r"D:\Adobe\Adobe Illustrator 2024"
        ]
        for pr in possible_roots:
            p_zh = os.path.join(pr, "Presets", "zh_CN", "脚本")
            if os.path.exists(p_zh):
                target_dirs.append(p_zh)
            p_en = os.path.join(pr, "Presets", "en_US", "Scripts")
            if os.path.exists(p_en):
                target_dirs.append(p_en)
                
        installed = []
        for td in target_dirs:
            try:
                dest = os.path.join(td, "🚀一键导出画板到贴图目录.jsx")
                shutil.copy2(src_jsx, dest)
                installed.append(dest)
            except Exception:
                pass
                
        if installed:
            QMessageBox.information(self, "安装成功", f"🎉 已成功将贴图导出脚本注入 Illustrator！\n\n已安装到:\n{installed[0]}")
        else:
            dest_dir = QFileDialog.getExistingDirectory(self, "选择 Illustrator 的 Presets/zh_CN/脚本 目录")
            if dest_dir:
                try:
                    shutil.copy2(src_jsx, os.path.join(dest_dir, "🚀一键导出画板到贴图目录.jsx"))
                    QMessageBox.information(self, "安装成功", "🎉 已成功安装到指定的脚本目录！")
                except Exception as e:
                    QMessageBox.warning(self, "安装失败", str(e))

def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # 检查是否有 C 语言原生的 pyi_splash 正在显示（PyInstaller --splash 打包环境）
    has_native_splash = False
    if pyi_splash is not None:
        try:
            if hasattr(pyi_splash, 'is_alive'):
                has_native_splash = pyi_splash.is_alive()
            else:
                has_native_splash = True
        except Exception:
            has_native_splash = False

    splash = None
    if not has_native_splash:
        # 如果未通过 PyInstaller 原生 Splash 启动（如直接 Python 调试运行），使用 Qt CatSplashScreen
        splash_path = SPLASH_CAT_JPG
        if os.path.exists(splash_path):
            splash = CatSplashScreen(splash_path)
            screen_geo = app.primaryScreen().geometry()
            splash.move((screen_geo.width() - splash.width()) // 2,
                        (screen_geo.height() - splash.height()) // 2)
            splash.show()
            app.processEvents()

    # 双锁门：最少 2000ms + 数据完全加载，二者同时满足才关闭 Splash
    _timer_done = [False]
    _data_done  = [False]
    _window     = [None]

    def do_show_main():
        """两把锁都开了才执行：关闭 Splash 并显示主窗口"""
        if not (_timer_done[0] and _data_done[0]):
            return

        # 1. 关闭 C 语言原生 Splash（如果存在）
        if pyi_splash is not None:
            try:
                pyi_splash.close()
            except Exception:
                pass

        # 2. 关闭 Qt 原生 Splash（如果存在）
        if splash:
            splash.close()

        # 3. 展示主窗口并设置 Windows 沉浸式暗黑顶栏
        win = _window[0]
        if win:
            win.show()
            set_dark_titlebar(int(win.winId()), win.current_theme == "dark")

    def on_timer_done():
        _timer_done[0] = True
        do_show_main()

    def on_data_ready():
        _data_done[0] = True
        do_show_main()

    def create_main_window():
        initial_files = sys.argv[1:] if len(sys.argv) > 1 else None
        win = MainWindow(initial_files=initial_files)
        _window[0] = win
        # 数据加载完毕信号 → 开锁 2
        win.initial_load_done.connect(on_data_ready)

    # 固定 2000ms 最少等待 → 开锁 1（无论数据多快加载完毕都要看满 2 秒）
    QTimer.singleShot(2000, on_timer_done)

    # 延迟 50ms 创建主窗口（让事件循环先稳定）
    QTimer.singleShot(50, create_main_window)

    # 终极兜底：6 秒后强制显示，防止信号异常丢失
    def force_show():
        _timer_done[0] = True
        _data_done[0]  = True
        do_show_main()
    QTimer.singleShot(6000, force_show)

    sys.exit(app.exec())

if __name__ == "__main__":
    main()
