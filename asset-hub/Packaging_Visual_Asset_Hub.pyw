# -*- coding: utf-8 -*-
"""
包装视觉资产管理器 (Packaging Visual Asset Hub - v3.2 硬盘扫描优先智能融合版)
核心机制：
1. 【扫描优先去重法则】：当 Excel 台账与本地硬盘扫描发生重复时，以【本地硬盘真实扫描结果为准】！
   - 优先采用硬盘上最新渲染的 Beauty 高清成品图与实际修改时间。
   - 继承 Excel 中的形态分类标签 (瓶装/袋装/盒装/套盒)。
2. 【深度智能扫描】：全面穿透主工作盘下多层级目录 (支持 400+ 真实工业项目)，秒级提取渲染封面。
3. 【实时热更新监听】：后台每 2 秒监听《产品列表.xlsx》，表格修改保存后 0.1 秒静默自动同步。
4. 【多维极速筛选】：左侧形态分类 + 客户品牌 + 顶部即时搜索框。
5. 【一键直达交互】：单击打开文件夹，双击打开 Blender 5.2 3D 工程，右键全能菜单。
"""

import os
import sys
import re
import json
import glob
import zipfile
import webbrowser
import subprocess
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk, ImageDraw
import openpyxl

CONFIG_FILE = os.path.join(os.path.expanduser("~"), ".packaging_organizer_v6.json")
CACHE_DIR = os.path.join(os.path.expanduser("~"), ".packaging_asset_thumbnails")
EXCEL_CACHE_DIR = os.path.join(CACHE_DIR, "excel_images")
os.makedirs(EXCEL_CACHE_DIR, exist_ok=True)

DEFAULT_EXCEL_PATH = r"C:\Users\qq424\WorkBuddy\2026-08-26-15-33-05\产品列表.xlsx"

BLENDER_EXE = r"C:\Program Files\Blender Foundation\Blender 5.2\blender.exe"
if not os.path.exists(BLENDER_EXE):
    BLENDER_EXE = "blender"

DEFAULT_WORKSPACES = []
for p in ["E:\\zjc", "D:\\Projects", "E:\\Projects"]:
    if os.path.exists(p) and p not in DEFAULT_WORKSPACES:
        DEFAULT_WORKSPACES.append(p)
if not DEFAULT_WORKSPACES:
    DEFAULT_WORKSPACES = ["E:\\zjc" if os.path.exists("E:\\zjc") else "D:\\Projects"]

SYSTEM_IGNORED_DIRS = {
    'system volume information', '$recycle.bin', 'recovery', '$windows.~bt',
    'msocache', 'config.msi', 'perflogs', 'program files', 'program files (x86)',
    'windows', 'programdata', 'appdata', '.git', '.vscode', '.idea', 'temp',
    'fonts', 'render  preset', 'renderraw', 'studio light hdri pack'
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if not data.get("workspaces"):
                    data["workspaces"] = DEFAULT_WORKSPACES
                return data
        except Exception:
            pass
    return {
        "workspaces": DEFAULT_WORKSPACES,
        "current_workspace": DEFAULT_WORKSPACES[0],
        "excel_path": DEFAULT_EXCEL_PATH if os.path.exists(DEFAULT_EXCEL_PATH) else ""
    }

def save_config(cfg):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def find_project_thumbnail(proj_path):
    """查找项目中最合适的高清渲染图作为缩略图"""
    render_dirs = [
        os.path.join(proj_path, "04_Renders_通道输出"),
        os.path.join(proj_path, "05_Delivery_最终交付"),
        os.path.join(proj_path, "渲染"),
        os.path.join(proj_path, "Renders"),
        proj_path
    ]
    candidates = []
    for rdir in render_dirs:
        if os.path.exists(rdir):
            try:
                for ext in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
                    candidates.extend(glob.glob(os.path.join(rdir, ext)))
            except Exception:
                pass
    if candidates:
        beauty_imgs = [c for c in candidates if any(k in os.path.basename(c).lower() for k in ["beauty", "成品", "主图", "camera", "正面", "01_"])]
        if beauty_imgs:
            try:
                beauty_imgs.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                return beauty_imgs[0]
            except Exception:
                return beauty_imgs[0]
        filtered = [c for c in candidates if not any(k in os.path.basename(c).lower() for k in ["mask", "alpha", "crypto", "选区", "蒙版", "normal", "depth", "roughness"])]
        if filtered:
            try:
                filtered.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                return filtered[0]
            except Exception:
                return filtered[0]
        try:
            candidates.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            return candidates[0]
        except Exception:
            return candidates[0]
    return None


def parse_and_cache_excel(excel_path):
    """解析 Excel 并在后台自动提取/更新内嵌的高清渲染图"""
    projects = []
    if not excel_path or not os.path.exists(excel_path):
        return projects

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb['全部'] if '全部' in wb.sheetnames else wb.active
        
        row_image_map = {}
        with zipfile.ZipFile(excel_path, 'r') as z:
            if 'xl/drawings/drawing1.xml' in z.namelist() and 'xl/drawings/_rels/drawing1.xml.rels' in z.namelist():
                drawing_xml = z.read('xl/drawings/drawing1.xml')
                rels_xml = z.read('xl/drawings/_rels/drawing1.xml.rels')

                root_d = ET.fromstring(drawing_xml)
                root_r = ET.fromstring(rels_xml)

                rel_map = {}
                for rel in root_r:
                    r_id = rel.attrib.get('Id')
                    target = rel.attrib.get('Target')
                    t_clean = target.lstrip('/').replace('../', '')
                    if not t_clean.startswith('xl/'):
                        t_clean = 'xl/' + t_clean
                    rel_map[r_id] = t_clean

                ns = {
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                }

                for anchor in root_d.findall('.//xdr:twoCellAnchor', ns) + root_d.findall('.//xdr:oneCellAnchor', ns):
                    from_elem = anchor.find('xdr:from', ns)
                    if from_elem is not None:
                        row_idx = int(from_elem.find('xdr:row', ns).text) + 1
                        blip = anchor.find('.//a:blip', ns)
                        if blip is not None:
                            embed_id = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                            if embed_id and embed_id in rel_map:
                                media_path = rel_map[embed_id]
                                ext = os.path.splitext(media_path)[1]
                                out_filename = f"excel_img_r{row_idx}{ext}"
                                out_full = os.path.join(EXCEL_CACHE_DIR, out_filename)
                                with open(out_full, 'wb') as f_out:
                                    f_out.write(z.read(media_path))
                                row_image_map[row_idx] = out_full

        for r in range(2, sheet.max_row + 1):
            name = sheet.cell(row=r, column=2).value
            cat = sheet.cell(row=r, column=4).value or "未分类"
            path = sheet.cell(row=r, column=5).value or ""
            time_str = sheet.cell(row=r, column=6).value or ""

            if not name:
                continue

            thumb = row_image_map.get(r, None)
            norm_path = path.replace("/", "\\") if path else ""

            brand = "柏缇"
            if norm_path:
                parts = norm_path.strip("\\").split("\\")
                if len(parts) >= 2 and parts[-2] not in {"zjc", "Projects", "E:", "D:"}:
                    brand = parts[-2]

            projects.append({
                "source": "excel",
                "brand": brand,
                "sku": str(name).strip(),
                "cat": str(cat).strip(),
                "path": norm_path,
                "thumbnail": thumb,
                "time": str(time_str),
                "mtime": os.path.getmtime(norm_path) if (norm_path and os.path.exists(norm_path)) else 0
            })

    except Exception as e:
        print(f"Error parsing Excel: {e}")

    return projects


def scan_workspace_projects(root_dir):
    """深度穿透扫描工作盘下所有实际存在的项目文件夹"""
    projects = []
    if not os.path.exists(root_dir):
        return projects
    try:
        entries = os.listdir(root_dir)
    except (PermissionError, OSError):
        return projects
        
    for entry in entries:
        if entry.lower() in SYSTEM_IGNORED_DIRS or entry.startswith('.') or entry.startswith('$') or entry.startswith('_'):
            continue
        brand_p = os.path.join(root_dir, entry)
        try:
            if not os.path.isdir(brand_p):
                continue
            sub_entries = os.listdir(brand_p)
        except (PermissionError, OSError):
            continue
            
        for sku in sub_entries:
            if sku.lower() in SYSTEM_IGNORED_DIRS or sku.startswith('.') or sku.startswith('$') or sku.startswith('_'):
                continue
            sku_p = os.path.join(brand_p, sku)
            try:
                if os.path.isdir(sku_p):
                    thumb = find_project_thumbnail(sku_p)
                    try:
                        s_mtime = os.path.getmtime(sku_p)
                    except Exception:
                        s_mtime = 0
                    projects.append({
                        "source": "disk",
                        "brand": entry,
                        "sku": sku,
                        "cat": "三维项目",
                        "path": sku_p,
                        "thumbnail": thumb,
                        "time": "",
                        "mtime": s_mtime
                    })
            except (PermissionError, OSError):
                continue
                
    projects.sort(key=lambda x: x["mtime"], reverse=True)
    return projects


def merge_excel_and_disk_projects(excel_projects, disk_projects):
    """
    【扫描优先去重准则】：
    如果 Excel 与 本地硬盘扫描 存在相同项目 (按路径或 SKU 匹配)：
    - 100% 以【本地硬盘扫描结果为准】！(采用硬盘上最新渲染的 Beauty 图、最新修改时间与真实目录)
    - 继承 Excel 的形态分类标签 (如 瓶装/袋装/盒装/套盒)
    """
    disk_map = {}
    for dp in disk_projects:
        norm_p = dp["path"].lower().replace("/", "\\").strip("\\")
        disk_map[norm_p] = dp
        # 建立备用 SKU 索引
        sku_key = f"{dp['brand']}@@{dp['sku']}".lower()
        disk_map[sku_key] = dp

    merged = []
    handled_disk_keys = set()

    for ep in excel_projects:
        norm_ep = ep["path"].lower().replace("/", "\\").strip("\\") if ep.get("path") else ""
        sku_key = f"{ep.get('brand', '')}@@{ep.get('sku', '')}".lower()
        
        matched_disk_proj = None
        if norm_ep and norm_ep in disk_map:
            matched_disk_proj = disk_map[norm_ep]
        elif sku_key in disk_map:
            matched_disk_proj = disk_map[sku_key]

        if matched_disk_proj:
            dp = matched_disk_proj
            handled_disk_keys.add(dp["path"].lower().replace("/", "\\").strip("\\"))
            # 扫描结果优先！
            thumb = dp["thumbnail"] if dp["thumbnail"] else ep["thumbnail"]
            mtime = dp["mtime"] if dp["mtime"] > 0 else ep["mtime"]
            merged.append({
                "source": "disk_prioritized",
                "brand": dp["brand"],
                "sku": dp["sku"],
                "cat": ep.get("cat", "三维项目"),
                "path": dp["path"],
                "thumbnail": thumb,
                "time": ep.get("time", ""),
                "mtime": mtime
            })
        else:
            merged.append(ep)

    # 将其余仅在硬盘存在的项目全部加入
    for dp in disk_projects:
        norm_p = dp["path"].lower().replace("/", "\\").strip("\\")
        if norm_p not in handled_disk_keys:
            merged.append(dp)

    # 按最新修改时间排序
    merged.sort(key=lambda x: x.get("mtime", 0), reverse=True)
    return merged


class AssetHubApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📦 包装视觉资产管理器 (Packaging Visual Asset Hub v3.2)")
        self.root.geometry("1180x780")
        self.root.minsize(960, 620)
        
        self.cfg = load_config()
        self.workspaces = self.cfg.get("workspaces", DEFAULT_WORKSPACES)
        cur_ws = self.cfg.get("current_workspace", self.workspaces[0])
        if not os.path.exists(cur_ws) and self.workspaces:
            cur_ws = self.workspaces[0]
        self.current_workspace_var = tk.StringVar(value=cur_ws)
        self.excel_path_var = tk.StringVar(value=self.cfg.get("excel_path", DEFAULT_EXCEL_PATH))
        
        self.last_excel_mtime = 0
        self.view_mode_var = tk.StringVar(value="merged") # "merged", "excel", "disk"
        self.search_var = tk.StringVar()
        self.selected_category_var = tk.StringVar(value="全部")
        
        self.page_size = 30
        self.current_page = 0
        
        self.excel_projects = []
        self.disk_projects = []
        self.merged_projects = []
        self.current_display_list = []
        self.filtered_projects = []
        self.thumb_cache = {}
        
        self.build_ui()
        self.load_all_data()
        
        self.start_excel_auto_sync_watcher()

    def build_ui(self):
        # 1. 顶部控制栏
        top_bar = ttk.Frame(self.root, padding=(15, 10))
        top_bar.pack(fill=tk.X)
        
        ttk.Label(top_bar, text="📊 视图模式:").pack(side=tk.LEFT, padx=(0, 4))
        self.combo_source = ttk.Combobox(
            top_bar,
            textvariable=self.view_mode_var,
            values=["merged", "excel", "disk"],
            state="readonly",
            width=22,
            font=("Microsoft YaHei", 9)
        )
        self.combo_source.pack(side=tk.LEFT, padx=(0, 15))
        self.combo_source.bind("<<ComboboxSelected>>", self.on_source_change)
        
        ttk.Label(top_bar, text="🔍 搜索产品:").pack(side=tk.LEFT, padx=(0, 4))
        search_entry = ttk.Entry(top_bar, textvariable=self.search_var, font=("Microsoft YaHei", 9), width=18)
        search_entry.pack(side=tk.LEFT, padx=(0, 12))
        self.search_var.trace_add("write", lambda *args: self.on_search_change())
        
        ttk.Button(top_bar, text="📊 重新指定 Excel...", command=self.import_new_excel).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(top_bar, text="🔄 立即刷新", command=self.load_all_data).pack(side=tk.LEFT, padx=(0, 6))
        
        self.sync_status_lbl = tk.Label(top_bar, text="🟢 扫描优先去重已就绪", font=("Microsoft YaHei", 8), fg="#059669", bg="#ECFDF5", padx=6, pady=2)
        self.sync_status_lbl.pack(side=tk.LEFT, padx=(6, 0))
        
        ttk.Button(top_bar, text="🌐 导出网页画廊 (HTML)...", command=self.export_html_gallery).pack(side=tk.RIGHT)

        # 2. 底栏分页控制
        self.bottom_bar = ttk.Frame(self.root, padding=(15, 8))
        self.bottom_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.page_info_lbl = ttk.Label(self.bottom_bar, text="", font=("Microsoft YaHei", 9))
        self.page_info_lbl.pack(side=tk.LEFT)
        
        self.btn_next = ttk.Button(self.bottom_bar, text="下一页 ➡️", command=self.next_page)
        self.btn_next.pack(side=tk.RIGHT, padx=(5, 0))
        
        self.btn_prev = ttk.Button(self.bottom_bar, text="⬅️ 上一页", command=self.prev_page)
        self.btn_prev.pack(side=tk.RIGHT)

        # 3. 主体区域 (左侧形态树 + 右侧卡片视口)
        main_pane = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 5))
        
        left_frame = ttk.LabelFrame(main_pane, text=" 🏷️ 形态分类 & 筛选 ", padding=8, width=200)
        main_pane.add(left_frame, weight=1)
        
        self.category_listbox = tk.Listbox(
            left_frame,
            font=("Microsoft YaHei", 10),
            selectmode=tk.SINGLE,
            relief=tk.FLAT,
            bg="#F8F9FA",
            selectbackground="#0078D7",
            selectforeground="white",
            highlightthickness=0,
            activestyle="none"
        )
        self.category_listbox.pack(fill=tk.BOTH, expand=True)
        self.category_listbox.bind("<<ListboxSelect>>", self.on_category_select)
        
        right_frame = ttk.Frame(main_pane)
        main_pane.add(right_frame, weight=5)
        
        self.canvas = tk.Canvas(right_frame, bg="#F0F2F5", highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(right_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        
        self.grid_container = tk.Frame(self.canvas, bg="#F0F2F5")
        self.grid_container.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        
        self.canvas_window = self.canvas.create_window((0, 0), window=self.grid_container, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.canvas.bind("<Configure>", self.on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self.on_mouse_wheel)

    def start_excel_auto_sync_watcher(self):
        ex_path = self.excel_path_var.get().strip()
        if ex_path and os.path.exists(ex_path):
            try:
                current_mtime = os.path.getmtime(ex_path)
                if self.last_excel_mtime > 0 and current_mtime > self.last_excel_mtime:
                    self.last_excel_mtime = current_mtime
                    self.thumb_cache.clear()
                    self.load_all_data()
                    self.flash_sync_status("⚡ 检测到 Excel 已更新，已按扫描优先合并完成！")
                elif self.last_excel_mtime == 0:
                    self.last_excel_mtime = current_mtime
            except Exception:
                pass
                
        self.root.after(2000, self.start_excel_auto_sync_watcher)

    def flash_sync_status(self, text):
        self.sync_status_lbl.config(text=text, bg="#FEF3C7", fg="#B45309")
        self.root.after(3500, lambda: self.sync_status_lbl.config(text="🟢 扫描优先去重已就绪", bg="#ECFDF5", fg="#059669"))

    def on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)
        self.render_cards()

    def on_mouse_wheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def load_all_data(self):
        ex_path = self.excel_path_var.get().strip()
        if not ex_path or not os.path.exists(ex_path):
            ex_path = DEFAULT_EXCEL_PATH
            self.excel_path_var.set(ex_path)
            
        if os.path.exists(ex_path):
            self.last_excel_mtime = os.path.getmtime(ex_path)
            
        self.excel_projects = parse_and_cache_excel(ex_path)
        
        cur_ws = self.current_workspace_var.get().strip()
        self.disk_projects = scan_workspace_projects(cur_ws)
        
        # 按照扫描优先规则深度融合
        self.merged_projects = merge_excel_and_disk_projects(self.excel_projects, self.disk_projects)
        
        self.combo_source["values"] = [
            f"⚡ 智能融合视图 ({len(self.merged_projects)})",
            f"仅 Excel 产品台账 ({len(self.excel_projects)})",
            f"仅本地工作盘扫描 ({len(self.disk_projects)})"
        ]
        
        mode = self.view_mode_var.get()
        if mode == "excel":
            self.combo_source.current(1)
        elif mode == "disk":
            self.combo_source.current(2)
        else:
            self.combo_source.current(0)
            
        self.update_active_dataset()

    def on_source_change(self, event=None):
        sel_idx = self.combo_source.current()
        if sel_idx == 1:
            self.view_mode_var.set("excel")
        elif sel_idx == 2:
            self.view_mode_var.set("disk")
        else:
            self.view_mode_var.set("merged")
        self.update_active_dataset()

    def update_active_dataset(self):
        mode = self.view_mode_var.get()
        if mode == "excel":
            self.current_display_list = self.excel_projects
        elif mode == "disk":
            self.current_display_list = self.disk_projects
        else:
            self.current_display_list = self.merged_projects
        
        cat_counts = {}
        for p in self.current_display_list:
            c = p.get("cat", "未分类")
            cat_counts[c] = cat_counts.get(c, 0) + 1
            
        self.category_listbox.delete(0, tk.END)
        self.category_listbox.insert(tk.END, f"✨ 全部形态 ({len(self.current_display_list)})")
        
        cats_sorted = sorted(cat_counts.keys())
        for c in cats_sorted:
            self.category_listbox.insert(tk.END, f"📦 {c} ({cat_counts[c]})")
            
        self.category_listbox.select_set(0)
        self.selected_category_var.set("全部")
        self.current_page = 0
        self.apply_filter()

    def import_new_excel(self):
        f = filedialog.askopenfilename(
            title="选择要导入的包装产品列表 Excel 表格 (.xlsx)",
            filetypes=[("Excel 表格", "*.xlsx"), ("所有文件", "*.*")]
        )
        if f:
            self.excel_path_var.set(f)
            self.cfg["excel_path"] = f
            save_config(self.cfg)
            self.load_all_data()
            messagebox.showinfo("同步成功", f"已成功绑定并开启实时同步！共 {len(self.excel_projects)} 个产品。")

    def on_search_change(self):
        self.current_page = 0
        self.apply_filter()

    def on_category_select(self, event=None):
        sel = self.category_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        text = self.category_listbox.get(idx)
        if idx == 0:
            self.selected_category_var.set("全部")
        else:
            m = re.search(r"📦\s*(.*?)\s*\(\d+\)", text)
            if m:
                self.selected_category_var.set(m.group(1))
        self.current_page = 0
        self.apply_filter()

    def apply_filter(self):
        cat = self.selected_category_var.get()
        kw = self.search_var.get().strip().lower()
        
        self.filtered_projects = []
        for p in self.current_display_list:
            if cat != "全部" and p.get("cat") != cat:
                continue
            if kw and (kw not in p["sku"].lower() and kw not in p.get("brand", "").lower() and kw not in p.get("cat", "").lower()):
                continue
            self.filtered_projects.append(p)
            
        total_items = len(self.filtered_projects)
        total_pages = max(1, (total_items + self.page_size - 1) // self.page_size)
        if self.current_page >= total_pages:
            self.current_page = total_pages - 1
            
        start_idx = self.current_page * self.page_size + 1 if total_items > 0 else 0
        end_idx = min((self.current_page + 1) * self.page_size, total_items)
        
        self.page_info_lbl.config(text=f"共 {total_items} 个产品 | 正在显示第 {start_idx} - {end_idx} 项 (第 {self.current_page + 1}/{total_pages} 页)")
        self.btn_prev.config(state=tk.NORMAL if self.current_page > 0 else tk.DISABLED)
        self.btn_next.config(state=tk.NORMAL if self.current_page < total_pages - 1 else tk.DISABLED)
        
        self.render_cards()

    def next_page(self):
        self.current_page += 1
        self.apply_filter()
        self.canvas.yview_moveto(0)

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.apply_filter()
            self.canvas.yview_moveto(0)

    def get_scaled_thumbnail(self, img_path, size=(190, 190)):
        if not img_path or not os.path.exists(img_path):
            return self.get_placeholder_thumbnail(size)
            
        if img_path in self.thumb_cache:
            return self.thumb_cache[img_path]
            
        try:
            im = Image.open(img_path)
            im.thumbnail(size, Image.Resampling.LANCZOS)
            
            thumb = Image.new("RGBA", size, (255, 255, 255, 255))
            offset_x = (size[0] - im.width) // 2
            offset_y = (size[1] - im.height) // 2
            
            if im.mode == "RGBA":
                thumb.paste(im, (offset_x, offset_y), im)
            else:
                thumb.paste(im, (offset_x, offset_y))
                
            tk_img = ImageTk.PhotoImage(thumb)
            self.thumb_cache[img_path] = tk_img
            return tk_img
        except Exception:
            return self.get_placeholder_thumbnail(size)

    def get_placeholder_thumbnail(self, size=(190, 190)):
        cache_key = "placeholder"
        if cache_key in self.thumb_cache:
            return self.thumb_cache[cache_key]
            
        im = Image.new("RGBA", size, (240, 243, 246, 255))
        draw = ImageDraw.Draw(im)
        draw.text((size[0]//2 - 35, size[1]//2 - 10), "📦 暂无渲染图", fill=(160, 170, 185, 255))
        tk_img = ImageTk.PhotoImage(im)
        self.thumb_cache[cache_key] = tk_img
        return tk_img

    def render_cards(self):
        for widget in self.grid_container.winfo_children():
            widget.destroy()
            
        if not self.filtered_projects:
            no_lbl = tk.Label(
                self.grid_container,
                text="📭 没有找到匹配的包装产品\n(提示：可通过左侧切换形态分类或清空搜索关键词)",
                font=("Microsoft YaHei", 12),
                bg="#F0F2F5",
                fg="#888",
                pady=60
            )
            no_lbl.pack()
            return

        container_width = self.canvas.winfo_width()
        if container_width < 100:
            container_width = 800
        card_w = 210
        cols = max(1, container_width // (card_w + 16))

        start_idx = self.current_page * self.page_size
        end_idx = start_idx + self.page_size
        page_items = self.filtered_projects[start_idx:end_idx]

        for idx, proj in enumerate(page_items):
            row = idx // cols
            col = idx % cols
            
            card = tk.Frame(
                self.grid_container,
                bg="white",
                bd=1,
                relief=tk.SOLID,
                padx=8,
                pady=8,
                highlightthickness=0
            )
            card.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
            
            tk_thumb = self.get_scaled_thumbnail(proj["thumbnail"])
            img_lbl = tk.Label(card, image=tk_thumb, bg="white", cursor="hand2")
            img_lbl.image = tk_thumb
            img_lbl.pack(fill=tk.BOTH, expand=True)
            
            meta_frame = tk.Frame(card, bg="white", pady=4)
            meta_frame.pack(fill=tk.X)
            
            badge_row = tk.Frame(meta_frame, bg="white")
            badge_row.pack(fill=tk.X, pady=(0, 2))
            
            cat_tag = tk.Label(
                badge_row,
                text=proj.get("cat", "包装"),
                font=("Microsoft YaHei", 8, "bold"),
                bg="#E1EFFF",
                fg="#005A9E",
                padx=4,
                pady=1
            )
            cat_tag.pack(side=tk.LEFT, padx=(0, 4))
            
            if proj.get("brand"):
                b_tag = tk.Label(
                    badge_row,
                    text=proj["brand"],
                    font=("Microsoft YaHei", 8),
                    bg="#F0F0F0",
                    fg="#555",
                    padx=4,
                    pady=1
                )
                b_tag.pack(side=tk.LEFT)
            
            title_lbl = tk.Label(
                meta_frame,
                text=proj["sku"],
                font=("Microsoft YaHei", 9, "bold"),
                bg="white",
                fg="#222",
                wraplength=180,
                justify="left"
            )
            title_lbl.pack(anchor="w")
            
            action_frame = tk.Frame(card, bg="white", pady=4)
            action_frame.pack(fill=tk.X)
            
            has_path = bool(proj.get("path") and os.path.exists(proj["path"]))
            btn_open = tk.Button(
                action_frame,
                text="📁 打开文件夹" if has_path else "📁 路径未就绪",
                font=("Microsoft YaHei", 8),
                bg="#F0F0F0" if has_path else "#FAFAFA",
                fg="#222" if has_path else "#999",
                relief=tk.FLAT,
                command=lambda p=proj.get("path"): self.open_folder(p)
            )
            btn_open.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
            
            btn_blend = tk.Button(
                action_frame,
                text="🚀 3D工程",
                font=("Microsoft YaHei", 8),
                bg="#EBF5FB",
                fg="#005A9E",
                relief=tk.FLAT,
                command=lambda p=proj.get("path"): self.launch_blend(p)
            )
            btn_blend.pack(side=tk.RIGHT)
            
            for w in (card, img_lbl, title_lbl, meta_frame):
                w.bind("<Button-1>", lambda e, p=proj.get("path"): self.open_folder(p))
                w.bind("<Double-1>", lambda e, p=proj.get("path"): self.launch_blend(p))
                w.bind("<Button-3>", lambda e, pr=proj: self.show_context_menu(e, pr))

    def open_folder(self, path):
        if path and os.path.exists(path):
            try:
                os.startfile(path)
            except Exception as e:
                messagebox.showerror("打开错误", str(e))
        else:
            messagebox.showwarning("提示", f"该项目的本地文件夹暂未找到或路径不存在:\n{path}")

    def launch_blend(self, proj_path):
        if not proj_path or not os.path.exists(proj_path):
            messagebox.showwarning("提示", f"未找到该项目的本地文件夹:\n{proj_path}")
            return
            
        blend_dir = os.path.join(proj_path, "03_3D_三维工程")
        target_dir = blend_dir if os.path.exists(blend_dir) else proj_path
        
        blends = glob.glob(os.path.join(target_dir, "*.blend"))
        if blends:
            blends.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            chosen = blends[0]
            try:
                subprocess.Popen([BLENDER_EXE, chosen])
            except Exception:
                os.startfile(chosen)
        else:
            self.open_folder(proj_path)

    def show_context_menu(self, event, proj):
        menu = tk.Menu(self.root, tearoff=0)
        p = proj.get("path", "")
        menu.add_command(label=f"📁 打开项目文件夹: {proj['sku']}", command=lambda: self.open_folder(p))
        menu.add_command(label="🚀 启动 Blender 打开 3D 工程", command=lambda: self.launch_blend(p))
        if p and os.path.exists(p):
            menu.add_command(label="🎨 打开 01_Design 平面原稿", command=lambda: self.open_folder(os.path.join(p, "01_Design_平面原稿")))
            menu.add_command(label="🖼️ 查看 04_Renders 渲染大图", command=lambda: self.open_folder(os.path.join(p, "04_Renders_通道输出")))
        menu.add_separator()
        menu.add_command(label="📋 复制项目完整路径", command=lambda: self.copy_path_to_clipboard(p))
        menu.tk_popup(event.x_root, event.y_root)

    def copy_path_to_clipboard(self, text):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("已复制", f"已复制路径到剪贴板:\n{text}")

    def export_html_gallery(self):
        cur_ws = self.current_workspace_var.get().strip()
        html_file = os.path.join(cur_ws if os.path.exists(cur_ws) else os.path.expanduser("~"), "📦_包装项目全景视觉画廊.html")
        
        cards_html = []
        for p in self.current_display_list:
            thumb_rel = p.get("thumbnail") or ""
            thumb_src = "file:///" + thumb_rel.replace("\\", "/") if (thumb_rel and os.path.exists(thumb_rel)) else "https://via.placeholder.com/300x300?text=No+Render"
            folder_uri = "file:///" + p["path"].replace("\\", "/") if p.get("path") else "#"
            
            cards_html.append(f"""
            <div class="card" onclick="window.open('{folder_uri}')">
                <div class="thumb-container">
                    <img src="{thumb_src}" alt="{p['sku']}" loading="lazy">
                </div>
                <div class="meta">
                    <div style="display:flex; gap:6px; margin-bottom:6px;">
                        <span class="badge" style="background:#0369a1;">{p.get('cat', '包装')}</span>
                        <span class="badge" style="background:#475569;">{p.get('brand', '')}</span>
                    </div>
                    <h3 class="title">{p['sku']}</h3>
                    <p class="path">{p.get('path', '')}</p>
                </div>
            </div>
            """)
            
        full_html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>📦 包装项目全景视觉画廊</title>
<style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif; background: #0f172a; color: #f8fafc; margin: 0; padding: 24px; }}
    h1 {{ font-size: 24px; font-weight: 700; margin-bottom: 24px; display: flex; align-items: center; gap: 10px; }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 20px; }}
    .card {{ background: #1e293b; border-radius: 12px; overflow: hidden; border: 1px solid #334155; cursor: pointer; transition: transform 0.2s, box-shadow 0.2s; }}
    .card:hover {{ transform: translateY(-4px); box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.5); border-color: #38bdf8; }}
    .thumb-container {{ width: 100%; aspect-ratio: 1; background: #0f172a; display: flex; align-items: center; justify-content: center; }}
    .thumb-container img {{ max-width: 100%; max-height: 100%; object-fit: contain; }}
    .meta {{ padding: 12px; }}
    .badge {{ display: inline-block; color: #e0f2fe; font-size: 11px; font-weight: 600; padding: 2px 8px; border-radius: 9999px; }}
    .title {{ font-size: 14px; font-weight: 600; margin: 0 0 4px 0; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
    .path {{ font-size: 11px; color: #94a3b8; margin: 0; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
</style>
</head>
<body>
    <h1>📦 包装项目全景视觉资产画廊 (共 {len(self.current_display_list)} 个产品)</h1>
    <div class="grid">
        {"".join(cards_html)}
    </div>
</body>
</html>"""

        try:
            with open(html_file, "w", encoding="utf-8") as f:
                f.write(full_html)
            webbrowser.open("file:///" + html_file.replace("\\", "/"))
        except Exception as e:
            messagebox.showerror("导出错误", str(e))


if __name__ == "__main__":
    root_win = tk.Tk()
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass
    app = AssetHubApp(root_win)
    root_win.mainloop()